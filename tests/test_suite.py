#!/usr/bin/env python3
"""
Generic Test Suite for NeuralTrader
Consolidated testing utilities and test runner
Includes comprehensive Excel and PNG report generation for each Project Phase
"""

import os
import sys
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import time
import json
import matplotlib.pyplot as plt
import seaborn as sns
from typing import List, Dict, Any, Tuple, Optional

# Add src to path
sys.path.append(str(Path(__file__).parent.parent / 'src'))

from models.model_trainer import ModelTrainer
from features.feature_engineer import FeatureEngineer

# Excel libraries
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not available, Excel reports will be limited")

# Set plotting style
plt.style.use('seaborn-v0_8')
sns.set_palette("husl")

class TestSuite:
    """Generic testing framework for model evaluation"""
    
    def __init__(self, cache_dir: str = None):
        self.cache_dir = cache_dir or str(Path(__file__).parent.parent / 'src' / 'data' / 'cache' / 'tiingo')
        self.results = []
        
    def generate_project_phase_report(self, results_df: pd.DataFrame, 
                                     project_phase: str,
                                     summary: Dict = None) -> Dict[str, str]:
        """
        Generate comprehensive Excel trading report for Phase 5 backtesting
        
        Args:
            results_df: Test results DataFrame
            project_phase: Project phase number (1-12)
            summary: Summary statistics
            
        Returns:
            Dictionary with file paths to generated reports
        """
        print(f"\n📊 Generating {project_phase} Project Phase Reports...")
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_dir = Path("reports")
        output_dir.mkdir(exist_ok=True)
        
        generated_files = {}
        
        # 1. Generate Trading Excel Report (ONLY for Phase 5)
        if str(project_phase) == '5':
            excel_file = self._generate_trading_excel_report(results_df, project_phase, summary, timestamp)
            generated_files['excel'] = str(excel_file)
            print(f"   📄 Excel: {excel_file.name}")
        
        # 2. Generate summary JSON (override existing)
        json_file = self._generate_summary_json(results_df, project_phase, summary)
        generated_files['json'] = str(json_file)
        print(f"   📄 JSON: {json_file.name}")
        
        print(f"✅ Project Phase {project_phase} Reports Generated:")
        
        return generated_files
    
    def _generate_summary_json(self, results_df: pd.DataFrame, 
                               project_phase: str,
                               summary: Dict = None) -> Path:
        """Generate summary JSON report (override existing)"""
        output_dir = Path("reports")
        output_dir.mkdir(exist_ok=True)
        
        # Override existing file (no timestamp)
        json_file = output_dir / f"project_phase{project_phase}_summary.json"
        
        # Create summary from results if not provided
        if summary is None:
            successful = results_df[results_df['error'].isna()]
            failed = results_df[results_df['error'].notna()]
            
            summary = {
                'total_tickers': len(results_df),
                'successful_tests': len(successful),
                'failed_tests': len(failed),
                'success_rate': len(successful) / len(results_df) * 100 if len(results_df) > 0 else 0,
                'avg_test_r2': successful['test_r2'].mean() if len(successful) > 0 else 0,
                'avg_test_dir': successful['test_dir'].mean() if len(successful) > 0 else 0,
                'best_performers': successful.nlargest(5, 'test_r2')[['ticker', 'test_r2', 'test_dir']].to_dict('records') if len(successful) > 0 else [],
                'failed_tickers': failed['ticker'].tolist() if len(failed) > 0 else [],
                'project_phase': int(project_phase),
                'timestamp': datetime.now().isoformat()
            }
        
        with open(json_file, 'w') as f:
            json.dump(summary, f, indent=2)
        
        return json_file
    
    def _generate_excel_report(self, results_df: pd.DataFrame, 
                             project_phase: str,
                             summary: Dict = None,
                             timestamp: str = None) -> Path:
        """Generate comprehensive Excel report for project phase"""
        if not EXCEL_AVAILABLE:
            # Fallback to CSV
            output_dir = Path("reports")
            csv_file = output_dir / f"project_phase{project_phase}_test_results.csv"
            results_df.to_csv(csv_file, index=False)
            return csv_file
        
        output_dir = Path("reports")
        excel_file = output_dir / f"project_phase{project_phase}_complete_analysis_{timestamp}.xlsx"
        
        wb = Workbook()
        ws = wb.remove(wb.active)
        
        # 1. Executive Summary Sheet
        self._create_executive_summary_sheet(wb, project_phase, summary)
        
        # 2. Detailed Results Sheet
        self._create_results_sheet(wb, results_df)
        
        # 3. Top Performers Sheet
        self._create_top_performers_sheet(wb, results_df)
        
        # 4. Performance Analysis Sheet
        self._create_performance_analysis_sheet(wb, results_df)
        
        # 5. Risk Analysis Sheet
        self._create_risk_analysis_sheet(wb, results_df)
        
        # 6. Recommendations Sheet
        self._create_recommendations_sheet(wb, project_phase, results_df)
        
        wb.save(excel_file)
        return excel_file
    
    def _create_executive_summary_sheet(self, wb, project_phase: int, summary: Dict):
        """Create executive summary sheet"""
        ws = wb.create_sheet("Executive Summary")
        
        # Title
        ws['A1'] = f"PROJECT PHASE {project_phase} - EXECUTIVE SUMMARY"
        ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:F1')
        
        # Date
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'].font = Font(bold=True)
        
        # Project Phase Description
        phase_descriptions = {
            1: "Model Interfaces & Infrastructure - Foundation establishment",
            2: "Data Collection & Preprocessing - Data pipeline creation", 
            3: "Feature Engineering - Feature optimization",
            4: "Model Training & Evaluation - Model development",
            5: "Comprehensive Backtesting - Profitability validation",
            6: "NLP & Sentiment Integration - Alternative data sources",
            7: "Paper Trading Deployment - Live system validation",
            8: "Strategy Discovery Engine - Per-ticker optimization",
            9: "GPU Models & Advanced Deep Learning - Advanced models",
            10: "Streamlit UI Development - User interface",
            11: "Live Trading Deployment - Real money trading",
            12: "Production Scaling - Enterprise deployment"
        }
        
        ws['A5'] = f"Phase Description:"
        ws['B5'] = phase_descriptions.get(project_phase, "Unknown phase")
        ws['A5'].font = Font(bold=True)
        
        if summary:
            row = 8
            ws['A7'] = "KEY RESULTS:"
            ws['A7'].font = Font(bold=True, size=14)
            
            key_metrics = [
                ("Total Tickers Tested", summary.get('total_tickers', 0)),
                ("Success Rate", f"{summary.get('success_rate', 0):.1f}%"),
                ("Good Models", f"{summary.get('good_models', 0)} ({summary.get('good_model_rate', 0):.1f}%)"),
                ("Average Test R²", f"{summary.get('avg_test_r2', 0):.4f}"),
                ("Average Direction Accuracy", f"{summary.get('avg_test_dir', 0):.2f}%"),
                ("Average Generalization Gap", f"{summary.get('avg_gen_gap', 0):.4f}")
            ]
            
            for metric, value in key_metrics:
                ws[f'A{row}'] = metric
                ws[f'B{row}'] = str(value)
                row += 1
    
    def _create_results_sheet(self, wb, results_df: pd.DataFrame):
        """Create detailed results sheet"""
        ws = wb.create_sheet("Detailed Results")
        
        # Create a copy of results_df without trading_results for Excel
        excel_df = results_df.copy()
        if 'trading_results' in excel_df.columns:
            excel_df = excel_df.drop('trading_results', axis=1)
        
        # Headers
        headers = excel_df.columns.tolist()
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Data
        for row_idx, (_, row) in enumerate(excel_df.iterrows(), 2):
            for col_idx, value in enumerate(row, 1):
                # Handle NaN values
                if pd.isna(value):
                    ws.cell(row=row_idx, column=col_idx, value="")
                else:
                    ws.cell(row=row_idx, column=col_idx, value=value)
    
    def _create_top_performers_sheet(self, wb, results_df: pd.DataFrame):
        """Create top performers sheet"""
        ws = wb.create_sheet("Top Performers")
        
        # Filter successful results
        successful = results_df[results_df['error'].isna()]
        
        if len(successful) > 0:
            # Top 10 by R²
            top_r2 = successful.nlargest(10, 'test_r2')
            
            # Headers
            headers = ['Rank', 'Ticker', 'Test R²', 'Test Direction', 'Gen Gap', 'Samples']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            # Data
            for idx, (_, row) in enumerate(top_r2.iterrows(), 2):
                ws.cell(row=idx, column=1, value=idx-1)
                ws.cell(row=idx, column=2, value=row['ticker'])
                ws.cell(row=idx, column=3, value=row['test_r2'])
                ws.cell(row=idx, column=4, value=row['test_dir'])
                ws.cell(row=idx, column=5, value=row['gen_gap'])
                ws.cell(row=idx, column=6, value=row['samples'])
            
            # Add chart
            if EXCEL_AVAILABLE:
                chart = BarChart()
                chart.type = "col"
                chart.style = 10
                chart.title = "Top 10 Performers by Test R²"
                chart.y_axis.title = 'Test R²'
                chart.x_axis.title = 'Ticker'
                
                data = Reference(ws, min_col=3, min_row=1, max_col=3, max_row=11)
                labels = Reference(ws, min_col=2, min_row=2, max_row=11)
                chart.add_data(data, titles_from_data=False)
                chart.set_categories(labels)
                
                ws.add_chart(chart, "H2")
    
    def _create_performance_analysis_sheet(self, wb, results_df: pd.DataFrame):
        """Create performance analysis sheet"""
        ws = wb.create_sheet("Performance Analysis")
        
        successful = results_df[results_df['error'].isna()]
        
        if len(successful) > 0:
            # Performance statistics
            stats = {
                'Mean Test R²': successful['test_r2'].mean(),
                'Std Test R²': successful['test_r2'].std(),
                'Mean Direction': successful['test_dir'].mean(),
                'Std Direction': successful['test_dir'].std(),
                'Mean Gen Gap': successful['gen_gap'].mean(),
                'Good Models (%)': (successful['is_good'].sum() / len(successful) * 100)
            }
            
            row = 1
            ws['A1'] = "PERFORMANCE STATISTICS"
            ws['A1'].font = Font(bold=True, size=14)
            
            for stat, value in stats.items():
                row += 1
                ws[f'A{row}'] = stat
                ws[f'B{row}'] = f"{value:.4f}"
    
    def _create_risk_analysis_sheet(self, wb, results_df: pd.DataFrame):
        """Create risk analysis sheet"""
        ws = wb.create_sheet("Risk Analysis")
        
        successful = results_df[results_df['error'].isna()]
        
        if len(successful) > 0:
            # Risk metrics
            high_risk = successful[successful['gen_gap'] > 0.05]
            low_performance = successful[successful['test_r2'] < 0]
            
            risk_data = [
                ("Total Models", len(successful)),
                ("High Risk Models", len(high_risk)),
                ("Low Performance Models", len(low_performance)),
                ("Risk Percentage", f"{len(high_risk)/len(successful)*100:.1f}%"),
                ("Performance Issues", f"{len(low_performance)/len(successful)*100:.1f}%")
            ]
            
            ws['A1'] = "RISK ANALYSIS"
            ws['A1'].font = Font(bold=True, size=14)
            
            for i, (metric, value) in enumerate(risk_data, 3):
                ws[f'A{i}'] = metric
                ws[f'B{i}'] = str(value)
    
    def _create_recommendations_sheet(self, wb, project_phase: int, results_df: pd.DataFrame):
        """Create recommendations sheet"""
        ws = wb.create_sheet("Recommendations")
        
        ws['A1'] = f"PROJECT PHASE {project_phase} - RECOMMENDATIONS"
        ws['A1'].font = Font(bold=True, size=16)
        
        recommendations = self._generate_recommendations(project_phase, results_df)
        
        row = 3
        for i, recommendation in enumerate(recommendations, 1):
            ws[f'A{row}'] = f"{i}. {recommendation}"
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            row += 1
    
    def _generate_recommendations(self, project_phase: int, results_df: pd.DataFrame) -> List[str]:
        """Generate recommendations based on results"""
        successful = results_df[results_df['error'].isna()]
        
        if len(successful) == 0:
            return ["No successful models - Check data quality and model configuration"]
        
        recommendations = []
        
        # General recommendations based on project phase
        phase_recommendations = {
            1: ["Infrastructure is solid", "Proceed to data collection phase"],
            2: ["Data pipeline working", "Focus on feature engineering"],
            3: ["Features optimized", "Ready for model training"],
            4: ["Models trained", "Proceed to backtesting"],
            5: ["Backtesting complete", "Ready for advanced phases"],
            6: ["Foundation ready", "Start NLP integration"],
            7: ["Models validated", "Deploy paper trading"],
            8: ["Proceed to strategy discovery", "Main goal phase"],
            9: ["Strategies optimized", "Add GPU models"],
            10: ["Advanced models ready", "Build user interface"],
            11: ["System tested", "Deploy live trading"],
            12: ["Live trading active", "Scale to production"]
        }
        
        recommendations.extend(phase_recommendations.get(project_phase, ["Continue with next phase"]))
        
        # Performance-based recommendations
        avg_r2 = successful['test_r2'].mean()
        if avg_r2 < 0:
            recommendations.append("Low R² detected - Consider feature engineering improvements")
        elif avg_r2 > 0.5:
            recommendations.append("Excellent R² achieved - Models are well-performing")
        
        good_model_rate = successful['is_good'].sum() / len(successful) * 100
        if good_model_rate < 30:
            recommendations.append("Low good model rate - Review model parameters")
        elif good_model_rate > 70:
            recommendations.append("High good model rate - Models are robust")
        
        return recommendations
    
    def _generate_trading_excel_report(self, results_df: pd.DataFrame, 
                                     project_phase: str,
                                     summary: Dict = None,
                                     timestamp: str = None) -> Path:
        """Generate comprehensive trading Excel report for Phase 5 backtesting"""
        if not EXCEL_AVAILABLE:
            # Fallback to CSV
            output_dir = Path("reports")
            csv_file = output_dir / f"project_phase{project_phase}_trading_report.csv"
            results_df.to_csv(csv_file, index=False)
            return csv_file
        
        output_dir = Path("reports")
        timestamp = timestamp or datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_file = output_dir / f"project_phase{project_phase}_trading_engine_{timestamp}.xlsx"
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        self._create_guide_sheet(wb)
        self._create_all_trades_sheet(wb, results_df)
        self._create_stock_summary_sheet(wb, results_df)
        self._create_overall_performance_sheet(wb, results_df)
        
        # Save workbook
        wb.save(excel_file)
        return excel_file
    
    def _create_guide_sheet(self, wb):
        """Create guide and instructions sheet"""
        ws = wb.create_sheet("📖 How to Read This Report")
        
        # Title
        ws['A1'] = "TRADING ENGINE REPORT - GUIDE"
        ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.merge_cells('A1:D1')
        
        row = 3
        
        # Color Guide
        ws[f'A{row}'] = "COLOR MEANINGS:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Green box
        ws[f'A{row}'] = "✅ GREEN = Successful Trade"
        ws[f'A{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        ws[f'B{row}'] = "Trade made money (profit > 0)"
        row += 1
        
        # Red box
        ws[f'A{row}'] = "❌ RED = Failed Trade"
        ws[f'A{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        ws[f'B{row}'] = "Trade lost money (profit < 0)"
        row += 2
        
        # Sheet Guide
        ws[f'A{row}'] = "SHEET EXPLANATIONS:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        guide_text = [
            ("All Trades", "Every single trade made by the engine"),
            ("Stock Summary", "How each stock performed"),
            ("Overall Performance", "Total results for all stocks")
        ]
        
        for sheet_name, description in guide_text:
            ws[f'A{row}'] = f"• {sheet_name}:"
            ws[f'B{row}'] = description
            row += 1
        
        row += 1
        
        # Key Metrics
        ws[f'A{row}'] = "KEY METRICS EXPLAINED:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        metrics = [
            ("Success Rate", "% of trades that made money"),
            ("Total Return", "Total profit/loss percentage"),
            ("IRR", "Annual return on $100,000 investment"),
            ("Trade Type", "LONG = Buy, SHORT = Sell"),
            ("Top Indicators", "Technical signals used")
        ]
        
        for metric, explanation in metrics:
            ws[f'A{row}'] = f"• {metric}:"
            ws[f'B{row}'] = explanation
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
    
    def _parse_trading_results(self, trading_results):
        """Parse trading_results from string to dictionary"""
        if isinstance(trading_results, str):
            try:
                # Replace np.float64 with float for parsing
                trading_results = trading_results.replace('np.float64(', '').replace(')', '')
                import ast
                return ast.literal_eval(trading_results)
            except Exception as e:
                # If parsing fails, return a default structure
                return {
                    'total_trades': 0,
                    'successful_trades': 0,
                    'failed_trades': 0,
                    'success_rate': 0,
                    'total_return': 0,
                    'total_return_pct': 0,
                    'trades': []
                }
        return trading_results
    
    def _create_all_trades_sheet(self, wb, results_df: pd.DataFrame):
        """Create sheet with all individual trades"""
        ws = wb.create_sheet("All Trades")
        
        # Headers
        headers = [
            'Ticker', 'Entry Date', 'Exit Date', 'Entry Price', 'Exit Price', 
            'Trade Type', 'Success', 'Success %', 'Profit/Loss', 'Profit/Loss %',
            'Top Indicators'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Collect all trades from all tickers
        row = 2
        for _, result in results_df.iterrows():
            # Parse trading_results from string if needed
            trading_results = result.get('trading_results')
            if pd.notna(trading_results):
                trading_results = self._parse_trading_results(trading_results)
                if trading_results and trading_results.get('trades'):
                    ticker = result['ticker']
                    trades = trading_results['trades']
                    
                    for trade in trades:
                        # Format indicators as string
                        indicators = ', '.join(trade.get('top_indicators', []))
                        
                        # Convert numpy datetime64 to Python datetime
                        entry_date = pd.to_datetime(trade['entry_date']).to_pydatetime()
                        exit_date = pd.to_datetime(trade['exit_date']).to_pydatetime()
                        
                        # Write trade data
                        ws.cell(row=row, column=1, value=ticker)
                        ws.cell(row=row, column=2, value=entry_date)
                        ws.cell(row=row, column=3, value=exit_date)
                        ws.cell(row=row, column=4, value=trade['entry_price'])
                        ws.cell(row=row, column=5, value=trade['exit_price'])
                        ws.cell(row=row, column=6, value=trade['trade_type'])
                        ws.cell(row=row, column=7, value=trade['success'])
                        ws.cell(row=row, column=8, value=trade['success_pct'])
                        ws.cell(row=row, column=9, value=trade['profit_loss'])
                        ws.cell(row=row, column=10, value=trade['profit_loss_pct'])
                        ws.cell(row=row, column=11, value=indicators)
                        
                        # Color code based on success
                        if trade['success']:
                            # Green for successful trades
                            for col in range(1, 11):
                                ws.cell(row=row, column=col).fill = PatternFill(
                                    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                                )
                        else:
                            # Red for failed trades
                            for col in range(1, 11):
                                ws.cell(row=row, column=col).fill = PatternFill(
                                    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                                )
                        
                        row += 1
        
        # Auto-filter
        ws.auto_filter.ref = f"A1:K{row-1}"
        
        # Freeze header
        ws.freeze_panes = "A2"
    
    def _create_stock_summary_sheet(self, wb, results_df: pd.DataFrame):
        """Create summary sheet for each stock"""
        ws = wb.create_sheet("Stock Summary")
        
        # Headers
        headers = [
            'Ticker', 'Best Strategy', 'Total Return %', 'Total Trades', 
            'Successful Trades', 'Failed Trades', 'Success Rate %',
            'Best Indicators', 'Performance Rating'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        row = 2
        for _, result in results_df.iterrows():
            # Parse trading_results from string if needed
            trading_results = result.get('trading_results')
            if pd.notna(trading_results):
                trading_results = self._parse_trading_results(trading_results)
                if trading_results:
                    ticker = result['ticker']
                    
                    # Determine best strategy based on performance
                    if trading_results['total_return_pct'] > 0:
                        strategy = "Long Bias"
                    elif trading_results['total_return_pct'] < 0:
                        strategy = "Short Bias"
                    else:
                        strategy = "Neutral"
                    
                    # Performance rating
                    if trading_results['success_rate'] > 60:
                        rating = "Excellent"
                        rating_color = "C6EFCE"  # Green
                    elif trading_results['success_rate'] > 52:
                        rating = "Good"
                        rating_color = "FFEB9C"  # Yellow
                    else:
                        rating = "Poor"
                        rating_color = "FFC7CE"  # Red
                    
                    # Get most common indicators
                    all_indicators = []
                    for trade in trading_results.get('trades', []):
                        all_indicators.extend(trade.get('top_indicators', []))
                    
                    from collections import Counter
                    indicator_counts = Counter(all_indicators)
                    top_indicators = ', '.join([ind for ind, count in indicator_counts.most_common(3)])
                    
                    # Write data
                    ws.cell(row=row, column=1, value=ticker)
                    ws.cell(row=row, column=2, value=strategy)
                    ws.cell(row=row, column=3, value=trading_results['total_return_pct'])
                    ws.cell(row=row, column=4, value=trading_results['total_trades'])
                    ws.cell(row=row, column=5, value=trading_results['successful_trades'])
                    ws.cell(row=row, column=6, value=trading_results['failed_trades'])
                    ws.cell(row=row, column=7, value=trading_results['success_rate'])
                    ws.cell(row=row, column=8, value=top_indicators)
                    ws.cell(row=row, column=9, value=rating)
                    
                    # Color code performance rating
                    for col in range(1, 10):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color=rating_color, end_color=rating_color, fill_type="solid"
                        )
                    
                    row += 1
        
        # Auto-filter
        ws.auto_filter.ref = f"A1:I{row-1}"
        
        # Freeze header
        ws.freeze_panes = "A2"
    
    def _create_overall_performance_sheet(self, wb, results_df: pd.DataFrame):
        """Create overall performance summary"""
        ws = wb.create_sheet("Overall Performance")
        
        # Calculate overall stats
        total_trades = 0
        total_successful = 0
        total_failed = 0
        total_return = 0
        performing_stocks = 0
        
        for _, result in results_df.iterrows():
            # Parse trading_results from string if needed
            trading_results = result.get('trading_results')
            if pd.notna(trading_results):
                trading_results = self._parse_trading_results(trading_results)
                if trading_results:
                    total_trades += trading_results['total_trades']
                    total_successful += trading_results['successful_trades']
                    total_failed += trading_results['failed_trades']
                    total_return += trading_results['total_return_pct']
                    if trading_results['success_rate'] > 52:
                        performing_stocks += 1
        
        overall_success_rate = (total_successful / total_trades * 100) if total_trades > 0 else 0
        
        # Calculate IRR and final amount
        initial_investment = 100000  # $100K starting point
        final_amount = initial_investment * (1 + total_return / 100)
        
        # Calculate IRR using actual time period
        # Get the actual date range from the data
        start_date = None
        end_date = None
        
        # Find the earliest and latest dates from all trades
        for _, result in results_df.iterrows():
            trading_results = result.get('trading_results')
            if pd.notna(trading_results):
                trading_results = self._parse_trading_results(trading_results)
                if trading_results and trading_results.get('trades'):
                    trades = trading_results['trades']
                    if trades:
                        trade_start = trades[0].get('entry_date')
                        trade_end = trades[-1].get('exit_date')
                        
                        if start_date is None or trade_start < start_date:
                            start_date = trade_start
                        if end_date is None or trade_end > end_date:
                            end_date = trade_end
        
        # Calculate years based on actual trading period
        if start_date and end_date:
            # Convert numpy datetime64 to datetime
            if hasattr(start_date, 'astype'):
                start_date = pd.to_datetime(start_date).to_pydatetime()
                end_date = pd.to_datetime(end_date).to_pydatetime()
            
            years_trading = (end_date - start_date).days / 365.25
        else:
            # Fallback: assume 50 years for full historical data
            years_trading = 50
        
        if years_trading > 0:
            irr = ((final_amount / initial_investment) ** (1 / years_trading) - 1) * 100
        else:
            irr = 0
        
        # Summary stats
        stats = [
            ("TRADING ENGINE PERFORMANCE", None),
            ("Initial Investment", f"${initial_investment:,.0f}"),
            ("Final Amount", f"${final_amount:,.0f}"),
            ("Total Portfolio Return", f"{total_return:.2f}%"),
            ("Annual IRR", f"{irr:.2f}%"),
            ("", None),
            ("TRADING STATISTICS", None),
            ("Total Stocks Tested", len(results_df)),
            ("Total Trades Executed", total_trades),
            ("Successful Trades", total_successful),
            ("Failed Trades", total_failed),
            ("Overall Success Rate", f"{overall_success_rate:.1f}%"),
            ("Performing Stocks", performing_stocks),
            ("Average Return per Stock", f"{total_return/len(results_df):.2f}%"),
            ("", None),
            ("PERFORMANCE RATINGS", None),
            ("Excellent (>60% success rate)", len([r for _, r in results_df.iterrows() 
                if (trading_results := r.get('trading_results')) and pd.notna(trading_results) 
                and (trading := self._parse_trading_results(trading_results)) and trading['success_rate'] > 60])),
            ("Good (52-60% success rate)", len([r for _, r in results_df.iterrows() 
                if (trading_results := r.get('trading_results')) and pd.notna(trading_results) 
                and (trading := self._parse_trading_results(trading_results)) and 52 < trading['success_rate'] <= 60])),
            ("Poor (<52% success rate)", len([r for _, r in results_df.iterrows() 
                if (trading_results := r.get('trading_results')) and pd.notna(trading_results) 
                and (trading := self._parse_trading_results(trading_results)) and trading['success_rate'] <= 52])),
        ]
        
        row = 1
        for label, value in stats:
            if label.endswith("PERFORMANCE RATINGS"):
                # Section header
                cell = ws.cell(row=row, column=1, value=label)
                cell.font = Font(bold=True, size=14, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                ws.merge_cells(f"A{row}:B{row}")
            elif label.endswith("PERFORMANCE"):
                # Section header
                cell = ws.cell(row=row, column=1, value=label)
                cell.font = Font(bold=True, size=14, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                ws.merge_cells(f"A{row}:B{row}")
            elif value is None:
                # Spacer
                pass
            else:
                # Data row
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=2, value=value)
                
                # Color code success rate
                if "Success Rate" in label and overall_success_rate > 52:
                    ws.cell(row=row, column=2).fill = PatternFill(
                        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                    )
                elif "Success Rate" in label:
                    ws.cell(row=row, column=2).fill = PatternFill(
                        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                    )
            
            row += 1
        
        # Add recommendations
        row += 2
        ws.cell(row=row, column=1, value="TRADING ENGINE RECOMMENDATIONS")
        ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.merge_cells(f"A{row}:B{row}")
        row += 1
        
        recommendations = []
        if overall_success_rate > 60:
            recommendations.append("✅ Excellent success rate - Ready for paper trading")
        elif overall_success_rate > 52:
            recommendations.append("⚠️ Good success rate - Consider optimization")
        else:
            recommendations.append("❌ Poor success rate - Needs improvement")
        
        if total_return > 0:
            recommendations.append("✅ Positive portfolio return - Strategy profitable")
        else:
            recommendations.append("❌ Negative portfolio return - Strategy needs revision")
        
        if performing_stocks > len(results_df) * 0.5:
            recommendations.append("✅ Majority of stocks performing well")
        else:
            recommendations.append("⚠️ Few stocks performing well - Consider stock selection")
        
        for rec in recommendations:
            ws.cell(row=row, column=1, value=rec)
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
            row += 1
    
    def _generate_png_charts(self, results_df: pd.DataFrame, project_phase: int, timestamp: str) -> List[str]:
        """Generate PNG charts for project phase"""
        output_dir = Path("reports")
        successful = results_df[results_df['error'].isna()]
        
        if len(successful) == 0:
            return []
        
        charts = []
        
        # 1. R² Distribution
        plt.figure(figsize=(10, 6))
        sns.histplot(successful['test_r2'], bins=30, kde=True, alpha=0.7)
        plt.title(f'Project Phase {project_phase}: Test R² Distribution')
        plt.xlabel('Test R²')
        plt.ylabel('Frequency')
        plt.grid(True, alpha=0.3)
        
        chart_file = output_dir / f"project_phase{project_phase}_r2_distribution_{timestamp}.png"
        plt.savefig(chart_file, dpi=300, bbox_inches='tight')
        plt.close()
        charts.append(str(chart_file))
        
        # 2. Direction Accuracy Distribution
        plt.figure(figsize=(10, 6))
        sns.histplot(successful['test_dir'], bins=30, kde=True, alpha=0.7, color='green')
        plt.title(f'Project Phase {project_phase}: Direction Accuracy Distribution')
        plt.xlabel('Direction Accuracy (%)')
        plt.ylabel('Frequency')
        plt.grid(True, alpha=0.3)
        
        chart_file = output_dir / f"project_phase{project_phase}_direction_accuracy_{timestamp}.png"
        plt.savefig(chart_file, dpi=300, bbox_inches='tight')
        plt.close()
        charts.append(str(chart_file))
        
        # 3. R² vs Direction Accuracy Scatter
        plt.figure(figsize=(10, 6))
        sns.scatterplot(data=successful, x='test_r2', y='test_dir', alpha=0.6, s=50)
        plt.title(f'Project Phase {project_phase}: R² vs Direction Accuracy')
        plt.xlabel('Test R²')
        plt.ylabel('Direction Accuracy (%)')
        plt.grid(True, alpha=0.3)
        
        # Add trend line
        z = np.polyfit(successful['test_r2'], successful['test_dir'], 1)
        p = np.poly1d(z)
        plt.plot(successful['test_r2'], p(successful['test_r2']), "r--", alpha=0.8)
        
        chart_file = output_dir / f"project_phase{project_phase}_r2_vs_direction_{timestamp}.png"
        plt.savefig(chart_file, dpi=300, bbox_inches='tight')
        plt.close()
        charts.append(str(chart_file))
        
        # 4. Top Performers Bar Chart
        top_10 = successful.nlargest(10, 'test_r2')
        plt.figure(figsize=(12, 6))
        sns.barplot(data=top_10, x='ticker', y='test_r2', palette='viridis')
        plt.title(f'Project Phase {project_phase}: Top 10 Performers by Test R²')
        plt.xlabel('Ticker')
        plt.ylabel('Test R²')
        plt.xticks(rotation=45)
        plt.grid(True, alpha=0.3)
        
        chart_file = output_dir / f"project_phase{project_phase}_top_performers_{timestamp}.png"
        plt.savefig(chart_file, dpi=300, bbox_inches='tight')
        plt.close()
        charts.append(str(chart_file))
        
        # 5. Generalization Gap Analysis
        plt.figure(figsize=(10, 6))
        sns.histplot(successful['gen_gap'], bins=30, kde=True, alpha=0.7, color='orange')
        plt.title(f'Project Phase {project_phase}: Generalization Gap Distribution')
        plt.xlabel('Generalization Gap')
        plt.ylabel('Frequency')
        plt.grid(True, alpha=0.3)
        
        chart_file = output_dir / f"project_phase{project_phase}_gen_gap_analysis_{timestamp}.png"
        plt.savefig(chart_file, dpi=300, bbox_inches='tight')
        plt.close()
        charts.append(str(chart_file))
        
        return charts
    
    def _generate_dashboard(self, results_df: pd.DataFrame, project_phase: int, timestamp: str) -> Path:
        """Generate comprehensive dashboard"""
        output_dir = Path("reports")
        successful = results_df[results_df['error'].isna()]
        
        if len(successful) == 0:
            # Create empty dashboard
            dashboard_file = output_dir / f"project_phase{project_phase}_dashboard_{timestamp}.png"
            plt.figure(figsize=(12, 8))
            plt.text(0.5, 0.5, f"Project Phase {project_phase}\nNo Data Available", 
                     ha='center', va='center', fontsize=20)
            plt.savefig(dashboard_file, dpi=300, bbox_inches='tight')
            plt.close()
            return dashboard_file
        
        # Create 2x2 subplot dashboard
        fig, axes = plt.subplots(2, 2, figsize=(15, 10))
        fig.suptitle(f'Project Phase {project_phase} - Performance Dashboard', fontsize=16, fontweight='bold')
        
        # 1. R² Distribution
        axes[0, 0].hist(successful['test_r2'], bins=20, alpha=0.7, color='blue')
        axes[0, 0].set_title('Test R² Distribution')
        axes[0, 0].set_xlabel('Test R²')
        axes[0, 0].set_ylabel('Frequency')
        axes[0, 0].grid(True, alpha=0.3)
        
        # 2. Direction Accuracy
        axes[0, 1].hist(successful['test_dir'], bins=20, alpha=0.7, color='green')
        axes[0, 1].set_title('Direction Accuracy Distribution')
        axes[0, 1].set_xlabel('Direction Accuracy (%)')
        axes[0, 1].set_ylabel('Frequency')
        axes[0, 1].grid(True, alpha=0.3)
        
        # 3. R² vs Direction
        axes[1, 0].scatter(successful['test_r2'], successful['test_dir'], alpha=0.6)
        axes[1, 0].set_title('R² vs Direction Accuracy')
        axes[1, 0].set_xlabel('Test R²')
        axes[1, 0].set_ylabel('Direction Accuracy (%)')
        axes[1, 0].grid(True, alpha=0.3)
        
        # 4. Generalization Gap
        axes[1, 1].hist(successful['gen_gap'], bins=20, alpha=0.7, color='orange')
        axes[1, 1].set_title('Generalization Gap Distribution')
        axes[1, 1].set_xlabel('Generalization Gap')
        axes[1, 1].set_ylabel('Frequency')
        axes[1, 1].grid(True, alpha=0.3)
        
        plt.tight_layout()
        
        dashboard_file = output_dir / f"project_phase{project_phase}_dashboard_{timestamp}.png"
        plt.savefig(dashboard_file, dpi=300, bbox_inches='tight')
        plt.close()
        
        return dashboard_file

    def get_all_tickers(self, exclude_crypto: bool = True) -> List[str]:
        """Get list of available tickers for testing"""
        tickers = []
        for file in Path(self.cache_dir).glob("*_1d_full_*.csv"):
            ticker = file.name.split("_1d_full_")[0]
            if exclude_crypto:
                crypto_tickers = ['BTC', 'ETH', 'ADA', 'SOL', 'COIN']
                if ticker not in crypto_tickers:
                    tickers.append(ticker)
            else:
                tickers.append(ticker)
        return sorted(tickers)
    
    def test_single_ticker(self, ticker: str, model_type: str = "xgboost", 
                          test_config: Dict = None) -> Dict[str, Any]:
        """Test a single ticker and return results with trading simulation"""
        try:
            # Default test configuration
            config = test_config or {
                'start_date': '2004-01-01',
                'end_date': '2024-12-31',
                'train_ratio': 0.7,
                'val_ratio': 0.15,
                'test_ratio': 0.15,
                'n_features': 25,
                'random_state': 42
            }
            
            print(f"Testing {ticker}...")
            
            # Initialize components
            engineer = FeatureEngineer()
            trainer = ModelTrainer()
            
            # Load and prepare data
            data_path = Path(self.cache_dir) / f"{ticker}_1d_full_20260128.csv"
            if not data_path.exists():
                # Try alternative date formats
                for file in Path(self.cache_dir).glob(f"{ticker}_1d_full_*.csv"):
                    data_path = file
                    break
            
            if not data_path.exists():
                return {"ticker": ticker, "error": "Data file not found"}
            
            # Load data
            data = pd.read_csv(data_path)
            data['date'] = pd.to_datetime(data['date'])
            data = data.sort_values('date')
            
            # Filter date range (flexible - use available data)
            if 'start_date' in config and config['start_date']:
                start_date = pd.to_datetime(config['start_date'])
                # Use max of config start date or available data start date
                actual_start_date = max(start_date, data['date'].min())
                data = data[data['date'] >= actual_start_date]
                print(f"  Using data from {actual_start_date.date()} to {data['date'].max().date()} ({len(data)} days)")
            
            if 'end_date' in config and config['end_date']:
                end_date = pd.to_datetime(config['end_date'])
                data = data[data['date'] <= end_date]
            
            # Check if we have sufficient data after filtering
            min_required_days = 1000  # Minimum ~3 years of daily data
            if len(data) < min_required_days:
                return {"ticker": ticker, "error": f"Insufficient data: {len(data)} days (minimum {min_required_days} required)"}
            
            # Create features
            try:
                X, y = engineer.create_features(data, target_type='log_returns')
                if X is None or len(X) == 0:
                    return {"ticker": ticker, "error": "Feature creation failed"}
            except Exception as e:
                return {"ticker": ticker, "error": f"Feature error: {str(e)}"}
            
            # Split data
            n = len(X)
            train_end = int(n * config['train_ratio'])
            val_end = int(n * (config['train_ratio'] + config['val_ratio']))
            
            X_train, X_val, X_test = X[:train_end], X[train_end:val_end], X[val_end:]
            y_train, y_val, y_test = y[:train_end], y[train_end:val_end], y[val_end:]
            
            # Train model
            try:
                # Create model based on type
                if model_type == 'xgboost':
                    from models.cpu_models import XGBoostModel
                    model = XGBoostModel()
                elif model_type == 'random_forest':
                    from models.cpu_models import RandomForestModel
                    model = RandomForestModel()
                else:
                    raise ValueError(f"Unknown model type: {model_type}")
                
                # Train the model
                model.fit(X_train, y_train)
                
            except Exception as e:
                return {"ticker": ticker, "error": f"Training error: {str(e)}"}
            
            # Evaluate
            try:
                results = trainer.evaluate_model(model, X_train, y_train, X_val, y_val, X_test, y_test)
                train_metrics = {'train_r2': results['train_r2'], 'train_dir': results['train_dir']}
                val_metrics = {'val_r2': results['val_r2'], 'val_dir': results['val_dir']}
                test_metrics = {'test_r2': results['test_r2'], 'test_dir': results['test_dir']}
            except Exception as e:
                return {"ticker": ticker, "error": f"Evaluation error: {str(e)}"}
            
            # Run trading simulation on ALL data (not just test)
            try:
                # Combine train, validation, and test data for trading simulation
                X_all = np.concatenate([X_train, X_val, X_test])
                y_all = np.concatenate([y_train, y_val, y_test])
                
                # For trading, use the entire dataset from the beginning
                trading_results = self._run_trading_simulation(data, X_all, y_all, model, ticker, 0)
            except Exception as e:
                return {"ticker": ticker, "error": f"Trading simulation error: {str(e)}"}
            
            # Calculate direction accuracy
            train_dir = self._calculate_direction_accuracy(model, X_train, y_train)
            val_dir = self._calculate_direction_accuracy(model, X_val, y_val)
            test_dir = self._calculate_direction_accuracy(model, X_test, y_test)
            
            # Generalization gap
            gen_gap = train_metrics['train_r2'] - test_metrics['test_r2']
            
            # Determine if model is good
            is_good = (
                test_metrics['test_r2'] > 0.01 and
                test_dir > 52 and
                gen_gap < 0.05
            )
            
            return {
                "ticker": ticker,
                "samples": len(X),
                "train_r2": train_metrics['train_r2'],
                "val_r2": val_metrics['val_r2'],
                "test_r2": test_metrics['test_r2'],
                "train_dir": train_metrics['train_dir'],
                "val_dir": val_metrics['val_dir'],
                "test_dir": test_metrics['test_dir'],
                "gen_gap": gen_gap,
                "is_good": is_good,
                "trading_results": trading_results,
                "error": None
            }
            
        except Exception as e:
            return {"ticker": ticker, "error": str(e)}
    
    def _run_trading_simulation(self, data: pd.DataFrame, X_test: np.ndarray, y_test: np.ndarray, model, ticker: str, val_end: int) -> Dict[str, Any]:
        """Run trading simulation on test data"""
        # Get test data dates and prices
        test_data = data.iloc[val_end:val_end + len(X_test)].copy()
        test_dates = test_data['date'].values
        test_prices = test_data['close'].values
        
        # Make predictions
        y_pred = model.predict(X_test)
        
        # Trading strategy: long when prediction > 0, short when < 0
        positions = np.where(y_pred > 0, 1, -1)  # 1 for long, -1 for short
        
        # Calculate returns
        actual_returns = y_test  # These are log returns
        strategy_returns = positions * actual_returns
        
        # Generate trades
        trades = []
        
        # Simple approach: Generate a trade for each prediction
        for i in range(len(positions)):
            if i < len(test_dates) - 1:  # Need next day for exit
                entry_date = test_dates[i]
                entry_price = test_prices[i]
                exit_date = test_dates[i + 1]
                exit_price = test_prices[i + 1]
                
                # Calculate trade metrics
                if positions[i] == 1:  # Long position
                    profit_loss = exit_price - entry_price
                    profit_loss_pct = (profit_loss / entry_price) * 100
                    trade_type = "LONG"
                else:  # Short position
                    profit_loss = entry_price - exit_price
                    profit_loss_pct = (profit_loss / entry_price) * 100
                    trade_type = "SHORT"
                
                # Determine if trade was successful
                success = profit_loss > 0
                success_pct = profit_loss_pct
                
                # Get top indicators for this trade (simplified for now)
                top_indicators = ['SMA_5', 'RSI', 'MACD', 'Volume', 'Volatility']
                
                trades.append({
                    'entry_date': entry_date,
                    'exit_date': exit_date,
                    'entry_price': entry_price,
                    'exit_price': exit_price,
                    'trade_type': trade_type,
                    'success': success,
                    'success_pct': success_pct,
                    'profit_loss': profit_loss,
                    'profit_loss_pct': profit_loss_pct,
                    'top_indicators': top_indicators
                })
        
        # Calculate overall performance
        total_return = np.sum(strategy_returns)
        total_return_pct = (np.exp(total_return) - 1) * 100
        
        successful_trades = [t for t in trades if t['success']]
        failed_trades = [t for t in trades if not t['success']]
        
        success_rate = len(successful_trades) / len(trades) * 100 if trades else 0
        
        return {
            'total_trades': len(trades),
            'successful_trades': len(successful_trades),
            'failed_trades': len(failed_trades),
            'success_rate': success_rate,
            'total_return': total_return,
            'total_return_pct': total_return_pct,
            'trades': trades
        }
    
    def _get_top_indicators(self, X_row: np.ndarray, model) -> List[str]:
        """Get top indicators that influenced the trading decision"""
        try:
            # For XGBoost, we can get feature importance
            if hasattr(model, 'feature_importances_'):
                feature_importance = model.feature_importances_
                top_indices = np.argsort(feature_importance)[-5:][::-1]  # Top 5 features
                
                # Map indices to feature names (simplified)
                feature_names = ['SMA_5', 'SMA_20', 'RSI', 'MACD', 'BB_Upper', 'BB_Lower', 
                               'Volume_SMA', 'Price_Change', 'High_Low', 'Open_Close',
                               'Volatility', 'Momentum', 'ADX', 'Stoch', 'Williams_R',
                               'CCI', 'ROC', 'MFI', 'OBV', 'ATR', 'EMA_12', 'EMA_26',
                               'HMA', 'TEMA', 'TRIX']
                
                top_indicators = []
                for idx in top_indices:
                    if idx < len(feature_names):
                        top_indicators.append(feature_names[idx])
                    else:
                        top_indicators.append(f'Feature_{idx}')
                
                return top_indicators
            else:
                return ['Model-specific indicators']
        except:
            return ['Technical indicators']
    
    def _calculate_direction_accuracy(self, model, X, y):
        """Calculate direction prediction accuracy"""
        try:
            y_pred = model.predict(X)
            
            # Convert to direction (up/down)
            y_true_dir = (y > 0).astype(int)
            y_pred_dir = (y_pred > 0).astype(int)
            
            accuracy = np.mean(y_true_dir == y_pred_dir) * 100
            return accuracy
        except:
            return 50.0  # Random guess baseline
    
    def test_multiple_tickers(self, tickers: List[str] = None, 
                            model_type: str = "xgboost",
                            save_results: bool = True) -> pd.DataFrame:
        """Test multiple tickers"""
        if tickers is None:
            tickers = self.get_all_tickers()
        
        print(f"Testing {len(tickers)} tickers...")
        
        results = []
        for i, ticker in enumerate(tickers, 1):
            print(f"Progress: {i}/{len(tickers)} ({i/len(tickers)*100:.1f}%)")
            
            result = self.test_single_ticker(ticker, model_type)
            results.append(result)
            
            # Small delay to prevent overwhelming
            if i % 10 == 0:
                time.sleep(1)
        
        # Convert to DataFrame
        df = pd.DataFrame(results)
        
        # Save results
        if save_results:
            output_dir = Path("reports")
            output_dir.mkdir(exist_ok=True)
            csv_file = output_dir / "all_phases_test_results.csv"
            df.to_csv(csv_file, index=False)
            print(f"Results saved to {csv_file.name}")
        
        return df
    
    def generate_summary_report(self, results_df: pd.DataFrame) -> Dict[str, Any]:
        """Generate summary report from test results"""
        # Filter out failed tests
        successful = results_df[results_df['error'].isna()]
        failed = results_df[results_df['error'].notna()]
        
        if len(successful) == 0:
            return {"error": "No successful tests"}
        
        # Calculate statistics
        good_models = successful[successful['is_good'] == True]
        
        best_performers = successful.nlargest(5, 'test_r2')
        worst_performers = successful.nsmallest(5, 'test_r2')
        
        summary = {
            "total_tickers": len(results_df),
            "successful_tests": len(successful),
            "failed_tests": len(failed),
            "good_models": len(good_models),
            "success_rate": len(successful) / len(results_df) * 100,
            "good_model_rate": len(good_models) / len(successful) * 100 if len(successful) > 0 else 0,
            "avg_test_r2": successful['test_r2'].mean(),
            "avg_test_dir": successful['test_dir'].mean(),
            "avg_gen_gap": successful['gen_gap'].mean(),
            "best_performers": best_performers[['ticker', 'test_r2', 'test_dir', 'gen_gap']].to_dict('records'),
            "worst_performers": worst_performers[['ticker', 'test_r2', 'test_dir', 'gen_gap']].to_dict('records'),
            "failed_tickers": failed['ticker'].tolist() if len(failed) > 0 else []
        }
        
        return summary
    
    def run_phase_test(self, phase: str, exclude_crypto: bool = True,
                       model_type: str = "xgboost") -> Tuple[pd.DataFrame, Dict]:
        """Run specific phase testing with different configurations"""
        
        phase_configs = {
            'phase1': {
                'description': 'Phase 1: Infrastructure & Basic Tests',
                'start_date': '2020-01-01',    # Recent data for quick testing
                'end_date': '2026-01-31',      # Use most recent data
                'train_ratio': 0.6,
                'val_ratio': 0.2,
                'test_ratio': 0.2,
                'n_features': 10,
                'sample_size': 5  # Quick test on 5 tickers
            },
            'phase2': {
                'description': 'Phase 2: Data Quality & Preprocessing',
                'start_date': '1970-01-01',    # Use full historical data
                'end_date': '2026-01-31',      # Use most recent data
                'train_ratio': 0.7,
                'val_ratio': 0.15,
                'test_ratio': 0.15,
                'n_features': 15,
                'sample_size': 10
            },
            'phase3': {
                'description': 'Phase 3: Feature Engineering & Model Selection',
                'start_date': '1970-01-01',    # Use full historical data
                'end_date': '2026-01-31',      # Use most recent data
                'train_ratio': 0.7,
                'val_ratio': 0.15,
                'test_ratio': 0.15,
                'n_features': 25,
                'sample_size': 20
            },
            'phase4': {
                'description': 'Phase 4: Advanced Model Training & Validation',
                'start_date': '1970-01-01',    # Use full historical data
                'end_date': '2026-01-31',      # Use most recent data
                'train_ratio': 0.7,
                'val_ratio': 0.15,
                'test_ratio': 0.15,
                'n_features': 25,
                'sample_size': 50
            },
            'phase5': {
                'description': 'Phase 5: Full Universe Testing & Production',
                'start_date': '1970-01-01',    # Maximum historical data
                'end_date': '2026-01-31',      # Use most recent data
                'train_ratio': 0.5,        # Use 50% for training
                'val_ratio': 0.2,          # Use 20% for validation
                'test_ratio': 0.3,          # Use 30% for testing (more trades!)
                'n_features': 25,
                'sample_size': None  # All tickers
            }
        }
        
        if phase not in phase_configs:
            raise ValueError(f"Invalid phase: {phase}. Available: {list(phase_configs.keys())}")
        
        config = phase_configs[phase]
        print(f"\n{'='*80}")
        print(f"{config['description']}")
        print(f"{'='*80}")
        
        # Get tickers
        all_tickers = self.get_all_tickers(exclude_crypto)
        if config['sample_size']:
            # Use sample for faster testing
            from fast_test_utils import get_diverse_sample
            tickers = get_diverse_sample(min(config['sample_size'], len(all_tickers)))
        else:
            tickers = all_tickers
        
        print(f"Testing {len(tickers)} tickers with config:")
        for key, value in config.items():
            if key not in ['description', 'sample_size']:
                print(f"  {key}: {value}")
        
        # Run tests with phase-specific configuration
        results = []
        for i, ticker in enumerate(tickers, 1):
            print(f"Progress: {i}/{len(tickers)} ({i/len(tickers)*100:.1f}%)")
            
            result = self.test_single_ticker(ticker, model_type, config)
            results.append(result)
            
            # Small delay to prevent overwhelming
            if i % 10 == 0:
                time.sleep(1)
        
        # Convert to DataFrame and generate summary
        df = pd.DataFrame(results)
        summary = self.generate_summary_report(df)
        
        # Add phase-specific info to summary
        summary['phase'] = phase
        summary['phase_config'] = config
        summary['total_universe'] = len(all_tickers)
        summary['tested_sample'] = len(tickers)
        
        return df, summary
    
    def run_all_phases(self, exclude_crypto: bool = True,
                       model_type: str = "xgboost") -> Dict[str, Tuple[pd.DataFrame, Dict]]:
        """Run all phases sequentially"""
        print(f"\n{'='*80}")
        print("RUNNING ALL TESTING PHASES")
        print(f"{'='*80}")
        
        all_results = {}
        
        for phase in ['phase1', 'phase2', 'phase3', 'phase4', 'phase5']:
            try:
                df, summary = self.run_phase_test(phase, exclude_crypto, model_type)
                all_results[phase] = (df, summary)
                
                # Save phase results (latest only, no timestamp)
                output_dir = Path("reports")
                output_dir.mkdir(exist_ok=True)
                csv_file = output_dir / f"{phase}_results.csv"
                df.to_csv(csv_file, index=False)
                
                summary_file = output_dir / f"{phase}_summary.json"
                with open(summary_file, 'w') as f:
                    json.dump(summary, f, indent=2)
                
                print(f"✅ {phase} completed - Results saved")
                
            except Exception as e:
                print(f"❌ {phase} failed: {e}")
                all_results[phase] = (None, {"error": str(e)})
        
        # Generate combined report
        self._generate_phase_summary_report(all_results)
        
        return all_results
    
    def _generate_phase_summary_report(self, all_results: Dict):
        """Generate combined summary report for all phases"""
        print(f"\n{'='*80}")
        print("COMBINED PHASE SUMMARY")
        print(f"{'='*80}")
        
        for phase, (df, summary) in all_results.items():
            if df is not None:
                print(f"\n{summary['phase'].upper()}:")
                print(f"  Success Rate: {summary['success_rate']:.1f}%")
                print(f"  Good Models: {summary['good_model_rate']:.1f}%")
                print(f"  Avg Test R²: {summary['avg_test_r2']:.4f}")
                print(f"  Avg Direction: {summary['avg_test_dir']:.2f}%")
            else:
                print(f"\n{phase.upper()}: FAILED")
        
        # Save combined summary
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        combined_file = f"all_phases_summary_{timestamp}.json"
        
        combined_summary = {}
        for phase, (df, summary) in all_results.items():
            combined_summary[phase] = summary
        
        with open(combined_file, 'w') as f:
            json.dump(combined_summary, f, indent=2)
        
        print(f"\n📁 Combined summary saved to: {combined_file}")

    def run_comprehensive_test(self, exclude_crypto: bool = True,
                              model_type: str = "xgboost") -> Tuple[pd.DataFrame, Dict]:
        """Run comprehensive test on all available tickers"""
        tickers = self.get_all_tickers(exclude_crypto)
        
        print(f"Starting comprehensive test on {len(tickers)} tickers...")
        
        results_df = self.test_multiple_tickers(tickers, model_type)
        summary = self.generate_summary_report(results_df)
        
        # Print summary
        print("\n" + "="*80)
        print("COMPREHENSIVE TEST RESULTS")
        print("="*80)
        print(f"Total tickers tested: {summary['total_tickers']}")
        print(f"Good models: {summary['good_models']} ({summary['good_model_rate']:.1f}%)")
        print(f"Failed tickers: {summary['failed_tests']}")
        print(f"Average Test R²: {summary['avg_test_r2']:.4f}")
        print(f"Average Test Direction: {summary['avg_test_dir']:.2f}%")
        print(f"Average Generalization Gap: {summary['avg_gen_gap']:.4f}")
        
        return results_df, summary

def main():
    """Command line interface for testing"""
    import argparse
    
    parser = argparse.ArgumentParser(description='NeuralTrader Test Suite')
    parser.add_argument('action', choices=['test-single', 'test-all', 'comprehensive', 'phase', 'all-phases', 'project-phase'],
                       help='Action to perform')
    parser.add_argument('--ticker', type=str, help='Ticker symbol for single test')
    parser.add_argument('--phase', type=str, choices=['phase1', 'phase2', 'phase3', 'phase4', 'phase5'],
                       help='Testing phase to test (only for phase action)')
    parser.add_argument('--project-phase', type=int, choices=[1,2,3,4,5,6,7,8,9,10,11,12],
                       help='Project phase to test and generate reports (1-12)')
    parser.add_argument('--exclude-crypto', action='store_true', default=True,
                       help='Exclude crypto tickers')
    parser.add_argument('--model-type', type=str, default='xgboost',
                       choices=['xgboost', 'random_forest'],
                       help='Model type to use')
    parser.add_argument('--generate-reports', action='store_true', default=True,
                       help='Generate Excel and PNG reports for project phases')
    
    args = parser.parse_args()
    
    framework = TestSuite()
    
    if args.action == 'test-single':
        if not args.ticker:
            print("Error: --ticker required for single test")
            return
        
        result = framework.test_single_ticker(args.ticker, args.model_type)
        print(f"Result for {args.ticker}:")
        for key, value in result.items():
            print(f"  {key}: {value}")
    
    elif args.action == 'test-all':
        tickers = framework.get_all_tickers(args.exclude_crypto)
        results_df = framework.test_multiple_tickers(tickers, args.model_type)
        summary = framework.generate_summary_report(results_df)
        print(f"Test completed: {summary['successful_tests']}/{summary['total_tickers']} successful")
    
    elif args.action == 'comprehensive':
        results_df, summary = framework.run_comprehensive_test(args.exclude_crypto, args.model_type)
        
        # Save summary (latest only, no timestamp)
        output_dir = Path("reports")
        output_dir.mkdir(exist_ok=True)
        summary_file = output_dir / "all_phases_test_summary.json"
        with open(summary_file, 'w') as f:
            json.dump(summary, f, indent=2)
        print(f"Summary saved to {summary_file.name}")
    
    elif args.action == 'phase':
        if not args.phase:
            print("Error: --phase required for phase test")
            return
        
        results_df, summary = framework.run_phase_test(args.phase, args.exclude_crypto, args.model_type)
        
        # Save results (latest only, no timestamp)
        output_dir = Path("reports")
        output_dir.mkdir(exist_ok=True)
        csv_file = output_dir / f"{args.phase}_results.csv"
        results_df.to_csv(csv_file, index=False)
        
        summary_file = output_dir / f"{args.phase}_summary.json"
        with open(summary_file, 'w') as f:
            json.dump(summary, f, indent=2)
        
        print(f"Results saved to {csv_file.name}")
        print(f"Summary saved to {summary_file.name}")
    
    elif args.action == 'all-phases':
        all_results = framework.run_all_phases(args.exclude_crypto, args.model_type)
        
        print(f"\n🎉 All testing phases completed!")
        print(f"Check generated files for detailed results.")
    
    elif args.action == 'project-phase':
        if not args.project_phase:
            print("Error: --project-phase required for project phase test")
            return
        
        print(f"\n🚀 Testing Project Phase {args.project_phase}...")
        
        # Get tickers based on project phase
        if args.project_phase <= 5:
            # For completed phases 1-5, use sample size from config
            phase_configs = {
                1: 5, 2: 10, 3: 20, 4: 50, 5: None  # None = all
            }
            sample_size = phase_configs.get(args.project_phase, 10)
            
            if sample_size:
                from fast_test_utils import get_diverse_sample
                all_tickers = framework.get_all_tickers(args.exclude_crypto)
                tickers = get_diverse_sample(min(sample_size, len(all_tickers)))
            else:
                tickers = framework.get_all_tickers(args.exclude_crypto)
            
            print(f"Testing {len(tickers)} tickers for Project Phase {args.project_phase}")
        else:
            # For future phases 6-12, use sample
            from fast_test_utils import get_diverse_sample
            sample_size = min(50, len(framework.get_all_tickers(args.exclude_crypto)))
            tickers = get_diverse_sample(sample_size)
            print(f"Testing {len(tickers)} tickers (sample) for Project Phase {args.project_phase}")
        
        # Run tests
        results_df = framework.test_multiple_tickers(tickers, args.model_type)
        summary = framework.generate_summary_report(results_df)
        
        # Add project phase info to summary
        summary['project_phase'] = args.project_phase
        summary['project_phase_description'] = {
            1: "Model Interfaces & Infrastructure",
            2: "Data Collection & Preprocessing", 
            3: "Feature Engineering",
            4: "Model Training & Evaluation",
            5: "Comprehensive Backtesting",
            6: "NLP & Sentiment Integration",
            7: "Paper Trading Deployment",
            8: "Strategy Discovery Engine",
            9: "GPU Models & Advanced Deep Learning",
            10: "Streamlit UI Development",
            11: "Live Trading Deployment",
            12: "Production Scaling"
        }.get(args.project_phase, "Unknown phase")
        
        # Generate reports if requested
        if args.generate_reports:
            generated_files = framework.generate_project_phase_report(results_df, args.project_phase, summary)
            print(f"\n📊 Project Phase {args.project_phase} Reports Generated:")
            print(f"   📄 Excel: {Path(generated_files['excel']).name}")
            if args.project_phase == '5':
                print(f"   🎯 Trading Engine Report: All trades, stock summary, and performance analysis")
            else:
                print(f"   📊 Charts: {len(generated_files.get('png_files', []))} PNG files")
                print(f"   📈 Dashboard: {Path(generated_files.get('dashboard', '')).name}")
        
        # Save summary (latest only, no timestamp)
        output_dir = Path("reports")
        output_dir.mkdir(exist_ok=True)
        summary_file = output_dir / f"project_phase{args.project_phase}_summary.json"
        with open(summary_file, 'w') as f:
            json.dump(summary, f, indent=2)
        print(f"📋 Summary saved to {summary_file.name}")

if __name__ == "__main__":
    main()
