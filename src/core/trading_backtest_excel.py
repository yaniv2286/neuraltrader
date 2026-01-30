"""
Trading Backtest Excel - Income-First, Fully Auditable
=======================================================
Follows the official specification for clear, trustworthy reporting.

Core Objective: Answer "If I invested $100,000, how much money did I make or lose?"

Sheet Order (STRICT):
1. How_To_Read (auto-open)
2. 100K_Reality_Check (main decision)
3. Equity_Curve
4. Trade_Stats
5. ALL_TRADES (full audit)
6. PER_STOCK_SUMMARY
7. STRATEGY_COMPARISON
"""
import pandas as pd
import numpy as np
from datetime import datetime
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference, BarChart


class Colors:
    """Color scheme for the report."""
    GREEN = "90EE90"      # 🟩 Good/Pass
    YELLOW = "FFFF99"     # 🟨 Warning
    RED = "FFB6C1"        # 🟥 Fail
    GREY = "D3D3D3"       # ⚪ Info
    WHITE = "FFFFFF"
    BLUE = "ADD8E6"
    HEADER = "4472C4"


class TradingBacktestExcel:
    """
    Creates income-first, fully auditable Excel reports.
    
    Philosophy: If it can't be understood quickly, it can't be trusted.
                If it can't be trusted, it can't be traded.
    """
    
    def __init__(self, results: Dict, config: Dict, output_path: str):
        self.results = results
        self.config = config
        self.output_path = output_path
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
        
        # Extract key metrics
        self.trades_df = results['trades']
        self.initial_capital = 100000
        self.final_capital = results['final_capital']
        self.total_profit = self.final_capital - self.initial_capital
        self.total_profit_pct = (self.total_profit / self.initial_capital) * 100
        self.cagr = results['cagr_pct']
        self.max_dd = results['max_drawdown_pct']
        self.years_tested = self._calculate_years()
        
    def _calculate_years(self) -> float:
        """Calculate years tested from trade dates."""
        if len(self.trades_df) == 0:
            return 0
        first_date = pd.to_datetime(self.trades_df['entry_date']).min()
        last_date = pd.to_datetime(self.trades_df['exit_date']).max()
        return (last_date - first_date).days / 365.25
    
    def _apply_color(self, cell, color: str):
        """Apply background color to cell."""
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    def _make_header(self, cell, text: str):
        """Style cell as header."""
        cell.value = text
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color=Colors.HEADER, end_color=Colors.HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _get_cagr_color(self, cagr: float) -> str:
        """Get color for CAGR value."""
        if cagr >= 25:
            return Colors.GREEN
        elif cagr >= 15:
            return Colors.YELLOW
        else:
            return Colors.RED
    
    def _get_dd_color(self, dd: float) -> str:
        """Get color for drawdown value (dd is negative)."""
        if dd >= -12:
            return Colors.GREEN
        elif dd >= -20:
            return Colors.YELLOW
        else:
            return Colors.RED
    
    def create_report(self):
        """Create all sheets in the correct order."""
        print("📊 Creating Trading Backtest Excel Report...")
        
        # Create sheets in STRICT order
        self._create_how_to_read()
        self._create_100k_reality_check()
        self._create_equity_curve()
        self._create_trade_stats()
        self._create_all_trades()
        self._create_per_stock_summary()
        self._create_strategy_comparison()
        
        # Save
        self.wb.save(self.output_path)
        print(f"   ✅ Saved to: {self.output_path}")
    
    def _create_how_to_read(self):
        """Sheet 1: How_To_Read (FIRST & AUTO-OPEN)"""
        ws = self.wb.create_sheet("How_To_Read", 0)
        
        row = 1
        
        # Purpose (Top, Large Font)
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = "PURPOSE"
        cell.font = Font(bold=True, size=16)
        row += 1
        
        ws.merge_cells(f'A{row}:F{row+1}')
        cell = ws[f'A{row}']
        cell.value = "This report shows what would have happened to real money.\nRead from top to bottom. No finance background required."
        cell.font = Font(size=12)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 3
        
        # How to Read (Step-by-Step)
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = "HOW TO READ (STEP-BY-STEP)"
        cell.font = Font(bold=True, size=14)
        row += 2
        
        steps = [
            "1. Start with 100K_Reality_Check",
            "2. Look at profit and worst drawdown",
            "3. If verdict is FAIL — stop",
            "4. If verdict is PASS — verify trades in ALL_TRADES",
            "5. Use PER_STOCK_SUMMARY to see what helps or hurts",
            "6. Use STRATEGY_COMPARISON only for future decisions"
        ]
        
        for step in steps:
            ws[f'A{row}'] = step
            ws[f'A{row}'].font = Font(size=11)
            row += 1
        
        row += 1
        
        # Color Legend
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = "COLOR LEGEND (MANDATORY)"
        cell.font = Font(bold=True, size=14)
        row += 2
        
        # Color table
        ws[f'A{row}'] = "Color"
        ws[f'B{row}'] = "Meaning"
        self._make_header(ws[f'A{row}'], "Color")
        self._make_header(ws[f'B{row}'], "Meaning")
        row += 1
        
        colors = [
            ("🟩 Green", "Good / Pass / Healthy", Colors.GREEN),
            ("🟨 Yellow", "Warning / Borderline", Colors.YELLOW),
            ("🟥 Red", "Fail / Unacceptable", Colors.RED),
            ("⚪ Grey", "Informational only", Colors.GREY)
        ]
        
        for emoji, meaning, color in colors:
            ws[f'A{row}'] = emoji
            ws[f'B{row}'] = meaning
            self._apply_color(ws[f'A{row}'], color)
            row += 1
        
        row += 1
        
        # Rule (Boxed & Bold)
        ws.merge_cells(f'A{row}:F{row+2}')
        cell = ws[f'A{row}']
        cell.value = "RULE:\n\nProfit without survivability is not a valid strategy.\nIf the drawdown is too deep, the system is NOT tradable."
        cell.font = Font(bold=True, size=12, color="FF0000")
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
    
    def _create_100k_reality_check(self):
        """Sheet 2: 100K_Reality_Check (MAIN DECISION SHEET)"""
        ws = self.wb.create_sheet("100K_Reality_Check")
        
        row = 1
        
        # Section A — Headline Summary
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "HEADLINE SUMMARY"
        cell.font = Font(bold=True, size=16)
        row += 2
        
        # Summary table
        summary_data = [
            ("Initial Investment", f"${self.initial_capital:,.0f}", Colors.GREY),
            ("Final Portfolio Value", f"${self.final_capital:,.0f}", Colors.GREEN if self.total_profit > 0 else Colors.RED),
            ("Total Profit ($)", f"${self.total_profit:,.0f}", Colors.GREEN if self.total_profit > 0 else Colors.RED),
            ("Total Profit (%)", f"{self.total_profit_pct:.2f}%", Colors.GREEN if self.total_profit_pct > 0 else Colors.RED),
            ("Years Tested", f"{self.years_tested:.1f}", Colors.GREY),
            ("CAGR", f"{self.cagr:.2f}%", self._get_cagr_color(self.cagr)),
            ("Max Drawdown", f"{self.max_dd:.2f}%", self._get_dd_color(self.max_dd))
        ]
        
        ws[f'A{row}'] = "Item"
        ws[f'B{row}'] = "Value"
        self._make_header(ws[f'A{row}'], "Item")
        self._make_header(ws[f'B{row}'], "Value")
        row += 1
        
        for item, value, color in summary_data:
            ws[f'A{row}'] = item
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(size=12)
            self._apply_color(ws[f'B{row}'], color)
            row += 1
        
        row += 2
        
        # Section B — Worst Moment
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "WORST MOMENT (REALITY CHECK)"
        cell.font = Font(bold=True, size=14)
        row += 2
        
        worst_value = self.initial_capital * (1 + self.max_dd / 100)
        dollar_dd = self.initial_capital - worst_value
        
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Value"
        self._make_header(ws[f'A{row}'], "Metric")
        self._make_header(ws[f'B{row}'], "Value")
        row += 1
        
        worst_data = [
            ("Worst Portfolio Value", f"${worst_value:,.0f}", Colors.RED if worst_value < 85000 else Colors.YELLOW),
            ("Dollar Drawdown", f"-${dollar_dd:,.0f}", Colors.RED if worst_value < 85000 else Colors.YELLOW),
            ("Max Drawdown %", f"{self.max_dd:.2f}%", self._get_dd_color(self.max_dd))
        ]
        
        for metric, value, color in worst_data:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            self._apply_color(ws[f'B{row}'], color)
            row += 1
        
        row += 2
        
        # Section C — FINAL VERDICT
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "FINAL VERDICT (BIG & CLEAR)"
        cell.font = Font(bold=True, size=16, color="FF0000")
        row += 2
        
        # Verdict logic
        capital_grew = self.total_profit > 0
        growth_strong = self.cagr >= 15
        dd_survivable = self.max_dd >= -20
        income_ready = capital_grew and growth_strong and dd_survivable
        
        verdict_data = [
            ("Did capital grow?", "YES" if capital_grew else "NO", Colors.GREEN if capital_grew else Colors.RED),
            ("Was growth strong enough?", "YES" if growth_strong else "NO", Colors.GREEN if growth_strong else Colors.RED),
            ("Was drawdown survivable?", "YES" if dd_survivable else "NO", Colors.GREEN if dd_survivable else Colors.RED),
            ("Income-Ready?", "PASS" if income_ready else "FAIL", Colors.GREEN if income_ready else Colors.RED)
        ]
        
        ws[f'A{row}'] = "Question"
        ws[f'B{row}'] = "Answer"
        self._make_header(ws[f'A{row}'], "Question")
        self._make_header(ws[f'B{row}'], "Answer")
        row += 1
        
        for question, answer, color in verdict_data:
            ws[f'A{row}'] = question
            ws[f'B{row}'] = answer
            ws[f'A{row}'].font = Font(size=12)
            ws[f'B{row}'].font = Font(bold=True, size=14)
            self._apply_color(ws[f'B{row}'], color)
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _create_equity_curve(self):
        """Sheet 3: Equity_Curve"""
        ws = self.wb.create_sheet("Equity_Curve")
        
        # Build equity curve from trades
        equity_data = []
        current_equity = self.initial_capital
        peak_equity = self.initial_capital
        
        for _, trade in self.trades_df.iterrows():
            date = trade['exit_date']
            pnl = trade['pnl']
            current_equity += pnl
            
            if current_equity > peak_equity:
                peak_equity = current_equity
            
            drawdown = ((current_equity - peak_equity) / peak_equity) * 100
            
            equity_data.append({
                'Date': date,
                'Equity': current_equity,
                'Drawdown_%': drawdown
            })
        
        equity_df = pd.DataFrame(equity_data)
        
        # Write to sheet
        ws['A1'] = "Date"
        ws['B1'] = "Equity ($)"
        ws['C1'] = "Drawdown (%)"
        self._make_header(ws['A1'], "Date")
        self._make_header(ws['B1'], "Equity ($)")
        self._make_header(ws['C1'], "Drawdown (%)")
        
        for idx, row_data in enumerate(equity_df.itertuples(), start=2):
            ws[f'A{idx}'] = row_data.Date
            ws[f'B{idx}'] = row_data.Equity
            ws[f'C{idx}'] = row_data.Drawdown_
        
        # Create chart
        chart = LineChart()
        chart.title = "Equity Curve"
        chart.y_axis.title = "Portfolio Value ($)"
        chart.x_axis.title = "Date"
        
        data = Reference(ws, min_col=2, min_row=1, max_row=len(equity_df)+1)
        chart.add_data(data, titles_from_data=True)
        
        ws.add_chart(chart, "E2")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    def _create_trade_stats(self):
        """Sheet 4: Trade_Stats (Simple Only)"""
        ws = self.wb.create_sheet("Trade_Stats")
        
        # Calculate stats
        total_trades = len(self.trades_df)
        winning_trades = self.trades_df[self.trades_df['pnl'] > 0]
        losing_trades = self.trades_df[self.trades_df['pnl'] < 0]
        
        avg_win_pct = winning_trades['return_pct'].mean() if len(winning_trades) > 0 else 0
        avg_loss_pct = losing_trades['return_pct'].mean() if len(losing_trades) > 0 else 0
        
        profit_factor = self.results.get('profit_factor', 0)
        expectancy = self.trades_df['pnl'].mean()
        
        # Write stats
        ws['A1'] = "Metric"
        ws['B1'] = "Value"
        self._make_header(ws['A1'], "Metric")
        self._make_header(ws['B1'], "Value")
        
        stats = [
            ("Total Trades", total_trades, Colors.GREY),
            ("Avg Win (%)", f"{avg_win_pct:.2f}%", Colors.GREEN),
            ("Avg Loss (%)", f"{avg_loss_pct:.2f}%", Colors.RED),
            ("Profit Factor", f"{profit_factor:.2f}", 
             Colors.GREEN if profit_factor >= 1.8 else Colors.YELLOW if profit_factor >= 1.5 else Colors.RED),
            ("Expectancy per Trade", f"${expectancy:.2f}", Colors.GREEN if expectancy > 0 else Colors.RED)
        ]
        
        row = 2
        for metric, value, color in stats:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            self._apply_color(ws[f'B{row}'], color)
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def _create_all_trades(self):
        """Sheet 5: ALL_TRADES (FULL AUDIT & EXPLAINABILITY)"""
        ws = self.wb.create_sheet("ALL_TRADES")
        
        # Define columns in REQUIRED ORDER
        columns = [
            # Identification
            "Trade_ID", "Symbol", "Strategy", "Direction",
            # Entry
            "Entry_Date", "Entry_Price", "Entry_Reason_Short", "Entry_Reason_Detail",
            # Exit
            "Exit_Date", "Exit_Price", "Exit_Reason_Short", "Exit_Reason_Detail",
            # Risk & Result
            "Position_Size_$", "Quantity", "Gross_PnL_$", "Gross_PnL_%",
            "Equity_After_Trade", "Drawdown_%"
        ]
        
        # Write headers
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            self._make_header(cell, col_name)
        
        # Write trade data
        for trade_idx, trade in enumerate(self.trades_df.iterrows(), start=2):
            _, trade_data = trade
            
            # Identification
            ws.cell(row=trade_idx, column=1, value=trade_idx-1)  # Trade_ID
            ws.cell(row=trade_idx, column=2, value=trade_data.get('ticker', ''))  # Symbol
            ws.cell(row=trade_idx, column=3, value='AI_Ensemble_V2')  # Strategy
            ws.cell(row=trade_idx, column=4, value=trade_data.get('direction', 'LONG'))  # Direction
            
            # Entry
            ws.cell(row=trade_idx, column=5, value=str(trade_data.get('entry_date', ''))[:10])
            ws.cell(row=trade_idx, column=6, value=trade_data.get('entry_price', 0))
            
            entry_reason = trade_data.get('entry_reason', '')
            entry_short = entry_reason.split('|')[0] if '|' in entry_reason else entry_reason[:50]
            ws.cell(row=trade_idx, column=7, value=entry_short)
            ws.cell(row=trade_idx, column=8, value=entry_reason)
            
            # Exit
            ws.cell(row=trade_idx, column=9, value=str(trade_data.get('exit_date', ''))[:10])
            ws.cell(row=trade_idx, column=10, value=trade_data.get('exit_price', 0))
            
            exit_reason = trade_data.get('exit_reason', '')
            exit_short = exit_reason.split('|')[0] if '|' in exit_reason else exit_reason[:50]
            ws.cell(row=trade_idx, column=11, value=exit_short)
            ws.cell(row=trade_idx, column=12, value=exit_reason)
            
            # Risk & Result
            position_size = trade_data.get('shares', 0) * trade_data.get('entry_price', 0)
            ws.cell(row=trade_idx, column=13, value=position_size)
            ws.cell(row=trade_idx, column=14, value=trade_data.get('shares', 0))
            
            pnl = trade_data.get('pnl', 0)
            pnl_pct = trade_data.get('return_pct', 0)
            ws.cell(row=trade_idx, column=15, value=pnl)
            ws.cell(row=trade_idx, column=16, value=pnl_pct)
            
            # Color PnL cells
            pnl_cell = ws.cell(row=trade_idx, column=15)
            pnl_pct_cell = ws.cell(row=trade_idx, column=16)
            color = Colors.GREEN if pnl > 0 else Colors.RED
            self._apply_color(pnl_cell, color)
            self._apply_color(pnl_pct_cell, color)
            
            ws.cell(row=trade_idx, column=17, value=trade_data.get('capital', 0))
            
            dd = trade_data.get('drawdown', 0) * 100
            dd_cell = ws.cell(row=trade_idx, column=18, value=dd)
            if dd < -20:
                self._apply_color(dd_cell, Colors.RED)
        
        # Adjust column widths
        for col_idx, col_name in enumerate(columns, start=1):
            if 'Reason_Detail' in col_name:
                ws.column_dimensions[chr(64+col_idx)].width = 60
            elif 'Reason_Short' in col_name:
                ws.column_dimensions[chr(64+col_idx)].width = 30
            elif col_name in ['Entry_Date', 'Exit_Date']:
                ws.column_dimensions[chr(64+col_idx)].width = 12
            else:
                ws.column_dimensions[chr(64+col_idx)].width = 15
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Enable filters
        ws.auto_filter.ref = f'A1:{chr(64+len(columns))}1'
    
    def _create_per_stock_summary(self):
        """Sheet 6: PER_STOCK_SUMMARY"""
        ws = self.wb.create_sheet("PER_STOCK_SUMMARY")
        
        # Group trades by ticker
        ticker_stats = []
        
        for ticker in self.trades_df['ticker'].unique():
            ticker_trades = self.trades_df[self.trades_df['ticker'] == ticker]
            
            total_pnl = ticker_trades['pnl'].sum()
            total_pnl_pct = (total_pnl / (ticker_trades['shares'] * ticker_trades['entry_price']).sum()) * 100 if len(ticker_trades) > 0 else 0
            avg_trade_pnl = ticker_trades['pnl'].mean()
            
            wins = ticker_trades[ticker_trades['pnl'] > 0]
            losses = ticker_trades[ticker_trades['pnl'] < 0]
            win_rate = (len(wins) / len(ticker_trades)) * 100 if len(ticker_trades) > 0 else 0
            
            avg_win = wins['pnl'].mean() if len(wins) > 0 else 0
            avg_loss = abs(losses['pnl'].mean()) if len(losses) > 0 else 1
            profit_factor = avg_win / avg_loss if avg_loss > 0 else 0
            
            max_dd = ticker_trades['drawdown'].min() * 100
            worst_trade = ticker_trades['pnl'].min()
            
            # Verdict
            if total_pnl > 0 and profit_factor >= 1.6 and max_dd >= -15:
                verdict = "KEEP"
                verdict_color = Colors.GREEN
            elif len(ticker_trades) < 5:
                verdict = "REVIEW"
                verdict_color = Colors.YELLOW
            else:
                verdict = "REMOVE"
                verdict_color = Colors.RED
            
            ticker_stats.append({
                'Symbol': ticker,
                'Trades_Count': len(ticker_trades),
                'Total_PnL_$': total_pnl,
                'Total_PnL_%': total_pnl_pct,
                'Avg_Trade_PnL_$': avg_trade_pnl,
                'Profit_Factor': profit_factor,
                'Win_Rate_%': win_rate,
                'Max_Drawdown_%': max_dd,
                'Worst_Trade_$': worst_trade,
                'Stock_Verdict': verdict,
                'Verdict_Color': verdict_color
            })
        
        # Sort by Total PnL descending
        ticker_stats = sorted(ticker_stats, key=lambda x: x['Total_PnL_$'], reverse=True)
        
        # Write headers
        headers = [
            "Symbol", "Trades_Count", "Total_PnL_$", "Total_PnL_%", "Avg_Trade_PnL_$",
            "Profit_Factor", "Win_Rate_%", "Max_Drawdown_%", "Worst_Trade_$", "Stock_Verdict"
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            self._make_header(cell, header)
        
        # Write data
        for row_idx, stats in enumerate(ticker_stats, start=2):
            ws.cell(row=row_idx, column=1, value=stats['Symbol'])
            ws.cell(row=row_idx, column=2, value=stats['Trades_Count'])
            ws.cell(row=row_idx, column=3, value=stats['Total_PnL_$'])
            ws.cell(row=row_idx, column=4, value=stats['Total_PnL_%'])
            ws.cell(row=row_idx, column=5, value=stats['Avg_Trade_PnL_$'])
            ws.cell(row=row_idx, column=6, value=stats['Profit_Factor'])
            ws.cell(row=row_idx, column=7, value=stats['Win_Rate_%'])
            ws.cell(row=row_idx, column=8, value=stats['Max_Drawdown_%'])
            ws.cell(row=row_idx, column=9, value=stats['Worst_Trade_$'])
            
            verdict_cell = ws.cell(row=row_idx, column=10, value=stats['Stock_Verdict'])
            self._apply_color(verdict_cell, stats['Verdict_Color'])
            verdict_cell.font = Font(bold=True)
        
        # Adjust column widths
        for col_idx in range(1, len(headers)+1):
            ws.column_dimensions[chr(64+col_idx)].width = 18
        
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{chr(64+len(headers))}1'
    
    def _create_strategy_comparison(self):
        """Sheet 7: STRATEGY_COMPARISON (FUTURE-FOCUSED)"""
        ws = self.wb.create_sheet("STRATEGY_COMPARISON")
        
        # For now, single strategy
        strategy_data = {
            'Strategy_ID': 'AI_Ensemble_V2',
            'Strategy_Type': 'ML Ensemble (XGBoost+RF+LGBM)',
            'Timeframe': f"{self.years_tested:.1f} years",
            'Market': 'US Stocks + ETFs',
            'CAGR_%': self.cagr,
            'Total_PnL_$': self.total_profit,
            'Total_PnL_%': self.total_profit_pct,
            'Max_Drawdown_%': self.max_dd,
            'Profit_Factor': self.results.get('profit_factor', 0),
            'Win_Rate_%': self.results.get('win_rate_pct', 0),
            'Trades_Per_Year': len(self.trades_df) / max(self.years_tested, 1),
            'Strategy_Status': 'PASS' if (self.cagr >= 15 and self.max_dd >= -20) else 'FAIL',
            'Capital_Eligibility': 'YES' if (self.cagr >= 15 and self.max_dd >= -20) else 'NO'
        }
        
        # Write headers
        headers = list(strategy_data.keys())
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            self._make_header(cell, header)
        
        # Write data
        for col_idx, (key, value) in enumerate(strategy_data.items(), start=1):
            cell = ws.cell(row=2, column=col_idx, value=value)
            
            # Color status cells
            if key == 'Strategy_Status':
                color = Colors.GREEN if value == 'PASS' else Colors.RED
                self._apply_color(cell, color)
                cell.font = Font(bold=True)
            elif key == 'Capital_Eligibility':
                color = Colors.GREEN if value == 'YES' else Colors.RED
                self._apply_color(cell, color)
                cell.font = Font(bold=True)
        
        # Adjust column widths
        for col_idx in range(1, len(headers)+1):
            ws.column_dimensions[chr(64+col_idx)].width = 20
        
        ws.freeze_panes = 'A2'


def create_trading_backtest_excel(results: Dict, config: Dict, output_path: str):
    """
    Create income-first, fully auditable Excel report.
    
    Args:
        results: Backtest results dictionary
        config: Strategy configuration
        output_path: Path to save Excel file
    """
    report = TradingBacktestExcel(results, config, output_path)
    report.create_report()
    return output_path
