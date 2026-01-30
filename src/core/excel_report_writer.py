"""
Generic Excel Report Writer for Trading Strategies
Strategy-agnostic design supporting 1 to N strategies.

6 Required Sheets:
1. Overview - Quick strategy assessment
2. Strategy_Comparison - Compare strategies against baseline
3. Trades - Full trade audit log
4. Per_Stock_Summary - Per-ticker analysis
5. Config_Snapshot - Reproducibility
6. How_To_Read - Documentation

Color Rules (rule-based, not manual):
- Green: good / improvement
- Yellow: neutral / acceptable  
- Red: bad / risk
- Gray: informational only
"""
import pandas as pd
import numpy as np
from typing import Dict, List, Optional, Any
from datetime import datetime
from dataclasses import dataclass, field
import warnings
warnings.filterwarnings('ignore')

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.formatting.rule import FormulaRule, CellIsRule
    from openpyxl.styles.differential import DifferentialStyle
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# Color definitions (hex codes)
class Colors:
    # Status colors
    GREEN = "00C851"
    YELLOW = "FFBB33"
    RED = "FF4444"
    GRAY = "AAAAAA"
    WHITE = "FFFFFF"
    
    # Dark versions for headers
    DARK_GREEN = "007E33"
    DARK_RED = "CC0000"
    
    # Column group backgrounds
    LIGHT_GRAY = "F5F5F5"
    LIGHT_BLUE = "E3F2FD"
    LIGHT_YELLOW = "FFF9C4"
    LIGHT_GREEN = "E8F5E9"


# Threshold definitions (fixed contract)
class Thresholds:
    # CAGR thresholds
    CAGR_GOOD = 25.0
    CAGR_ACCEPTABLE = 15.0
    
    # Max Drawdown thresholds (absolute values)
    DD_GOOD = 15.0
    DD_ACCEPTABLE = 25.0
    
    # Profit Factor thresholds
    PF_GOOD = 1.5
    PF_ACCEPTABLE = 1.1
    
    # Win Rate thresholds
    WR_GOOD = 55.0
    WR_ACCEPTABLE = 45.0
    
    # Comparison tolerance
    COMPARISON_TOLERANCE = 5.0  # ±5%


@dataclass
class StrategyResult:
    """Container for strategy backtest results."""
    strategy_id: str
    cagr_pct: float
    max_drawdown_pct: float
    profit_factor: float
    win_rate_pct: float
    total_trades: int
    best_year_pct: float = 0.0
    worst_year_pct: float = 0.0
    total_return_pct: float = 0.0
    trades: pd.DataFrame = field(default_factory=pd.DataFrame)
    config: Dict[str, Any] = field(default_factory=dict)
    
    @property
    def status(self) -> str:
        """Determine PASS/FAIL based on thresholds."""
        if (self.cagr_pct >= Thresholds.CAGR_ACCEPTABLE and
            abs(self.max_drawdown_pct) <= Thresholds.DD_ACCEPTABLE and
            self.profit_factor >= Thresholds.PF_ACCEPTABLE and
            self.win_rate_pct >= Thresholds.WR_ACCEPTABLE):
            return "PASS"
        return "FAIL"


class ExcelReportWriter:
    """
    Generic Excel report writer for trading strategies.
    Supports 1 to N strategies with consistent formatting.
    """
    
    def __init__(self):
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl required for Excel reports")
        
        self.wb = Workbook()
        self.strategies: List[StrategyResult] = []
        self.baseline_id: Optional[str] = None
        
    def add_strategy(self, result: StrategyResult, is_baseline: bool = False):
        """Add a strategy result to the report."""
        self.strategies.append(result)
        if is_baseline or self.baseline_id is None:
            self.baseline_id = result.strategy_id
    
    def _get_cagr_color(self, value: float) -> str:
        """Get color for CAGR value."""
        if value >= Thresholds.CAGR_GOOD:
            return Colors.GREEN
        elif value >= Thresholds.CAGR_ACCEPTABLE:
            return Colors.YELLOW
        return Colors.RED
    
    def _get_dd_color(self, value: float) -> str:
        """Get color for Max Drawdown value (use absolute)."""
        abs_value = abs(value)
        if abs_value <= Thresholds.DD_GOOD:
            return Colors.GREEN
        elif abs_value <= Thresholds.DD_ACCEPTABLE:
            return Colors.YELLOW
        return Colors.RED
    
    def _get_pf_color(self, value: float) -> str:
        """Get color for Profit Factor value."""
        if value >= Thresholds.PF_GOOD:
            return Colors.GREEN
        elif value >= Thresholds.PF_ACCEPTABLE:
            return Colors.YELLOW
        return Colors.RED
    
    def _get_wr_color(self, value: float) -> str:
        """Get color for Win Rate value."""
        if value >= Thresholds.WR_GOOD:
            return Colors.GREEN
        elif value >= Thresholds.WR_ACCEPTABLE:
            return Colors.YELLOW
        return Colors.RED
    
    def _get_comparison_color(self, current: float, baseline: float, higher_is_better: bool = True) -> str:
        """Get color for comparison (current vs baseline)."""
        if baseline == 0:
            return Colors.GRAY
        
        pct_diff = ((current - baseline) / abs(baseline)) * 100
        
        if higher_is_better:
            if pct_diff > Thresholds.COMPARISON_TOLERANCE:
                return Colors.GREEN
            elif pct_diff < -Thresholds.COMPARISON_TOLERANCE:
                return Colors.RED
        else:  # Lower is better (like drawdown)
            if pct_diff < -Thresholds.COMPARISON_TOLERANCE:
                return Colors.GREEN
            elif pct_diff > Thresholds.COMPARISON_TOLERANCE:
                return Colors.RED
        
        return Colors.YELLOW
    
    def _get_comparison_verdict(self, current: float, baseline: float, higher_is_better: bool = True) -> str:
        """Get verdict for comparison."""
        if current == baseline:
            return "Baseline" if len(self.strategies) == 1 else "Same"
        
        pct_diff = ((current - baseline) / abs(baseline)) * 100 if baseline != 0 else 0
        
        if abs(pct_diff) <= Thresholds.COMPARISON_TOLERANCE:
            return "Same"
        
        if higher_is_better:
            return "Better" if current > baseline else "Worse"
        else:
            return "Better" if current < baseline else "Worse"
    
    def _apply_fill(self, cell, color: str):
        """Apply fill color to cell."""
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    def _apply_header_style(self, cell):
        """Apply header style to cell."""
        cell.font = Font(bold=True, color=Colors.WHITE)
        cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _create_overview_sheet(self):
        """Create Sheet 1: Overview - Quick strategy assessment."""
        ws = self.wb.active
        ws.title = "Overview"
        
        # Headers
        headers = [
            "strategy_id",
            "Annual Growth % (CAGR)",
            "Worst Loss % (Max Drawdown)",
            "Profit / Loss Ratio",
            "Win Rate %",
            "Total Trades",
            "Best Year %",
            "Worst Year %",
            "Status"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)
        
        # Data rows
        for row_idx, strategy in enumerate(self.strategies, 2):
            ws.cell(row=row_idx, column=1, value=strategy.strategy_id)
            
            # CAGR with color
            cell = ws.cell(row=row_idx, column=2, value=round(strategy.cagr_pct, 2))
            self._apply_fill(cell, self._get_cagr_color(strategy.cagr_pct))
            
            # Max Drawdown with color
            cell = ws.cell(row=row_idx, column=3, value=round(strategy.max_drawdown_pct, 2))
            self._apply_fill(cell, self._get_dd_color(strategy.max_drawdown_pct))
            
            # Profit Factor with color
            cell = ws.cell(row=row_idx, column=4, value=round(strategy.profit_factor, 2))
            self._apply_fill(cell, self._get_pf_color(strategy.profit_factor))
            
            # Win Rate with color
            cell = ws.cell(row=row_idx, column=5, value=round(strategy.win_rate_pct, 2))
            self._apply_fill(cell, self._get_wr_color(strategy.win_rate_pct))
            
            # Total Trades (no color)
            ws.cell(row=row_idx, column=6, value=strategy.total_trades)
            
            # Best/Worst Year
            ws.cell(row=row_idx, column=7, value=round(strategy.best_year_pct, 2))
            ws.cell(row=row_idx, column=8, value=round(strategy.worst_year_pct, 2))
            
            # Status with color
            cell = ws.cell(row=row_idx, column=9, value=strategy.status)
            if strategy.status == "PASS":
                cell.fill = PatternFill(start_color=Colors.DARK_GREEN, end_color=Colors.DARK_GREEN, fill_type="solid")
                cell.font = Font(bold=True, color=Colors.WHITE)
            else:
                cell.fill = PatternFill(start_color=Colors.DARK_RED, end_color=Colors.DARK_RED, fill_type="solid")
                cell.font = Font(bold=True, color=Colors.WHITE)
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 20
    
    def _create_comparison_sheet(self):
        """Create Sheet 2: Strategy_Comparison."""
        ws = self.wb.create_sheet("Strategy_Comparison")
        
        # Get baseline
        baseline = next((s for s in self.strategies if s.strategy_id == self.baseline_id), self.strategies[0])
        
        # Headers
        headers = ["metric", "baseline_value", "current_strategy_value", "difference", "verdict"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)
        
        # Metrics to compare
        metrics = [
            ("Annual Growth % (CAGR)", "cagr_pct", True),
            ("Worst Loss % (Max Drawdown)", "max_drawdown_pct", False),  # Lower is better
            ("Profit / Loss Ratio", "profit_factor", True),
            ("Win Rate %", "win_rate_pct", True),
            ("Total Trades", "total_trades", True),
            ("Total Return %", "total_return_pct", True),
        ]
        
        row = 2
        for strategy in self.strategies:
            for metric_name, attr, higher_is_better in metrics:
                baseline_val = getattr(baseline, attr)
                current_val = getattr(strategy, attr)
                diff = current_val - baseline_val
                verdict = self._get_comparison_verdict(current_val, baseline_val, higher_is_better)
                
                ws.cell(row=row, column=1, value=f"{strategy.strategy_id}: {metric_name}")
                ws.cell(row=row, column=2, value=round(baseline_val, 2))
                ws.cell(row=row, column=3, value=round(current_val, 2))
                ws.cell(row=row, column=4, value=round(diff, 2))
                
                cell = ws.cell(row=row, column=5, value=verdict)
                if verdict == "Better":
                    self._apply_fill(cell, Colors.GREEN)
                elif verdict == "Worse":
                    self._apply_fill(cell, Colors.RED)
                else:
                    self._apply_fill(cell, Colors.YELLOW)
                
                row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        for col in 'BCDE':
            ws.column_dimensions[col].width = 20
    
    def _create_trades_sheet(self):
        """Create Sheet 3: Trades - Full audit log with detailed signal reasons."""
        ws = self.wb.create_sheet("Trades")
        
        # Headers with column groups - enhanced for detailed reasons
        headers = [
            ("strategy_id", Colors.LIGHT_GRAY, 15),
            ("ticker", Colors.LIGHT_GRAY, 8),
            ("direction", Colors.LIGHT_GRAY, 8),
            ("entry_date", Colors.LIGHT_BLUE, 12),
            ("exit_date", Colors.LIGHT_BLUE, 12),
            ("hold_days", Colors.LIGHT_BLUE, 10),
            ("entry_price", Colors.WHITE, 12),
            ("exit_price", Colors.WHITE, 12),
            ("shares", Colors.WHITE, 10),
            ("pnl $", None, 12),  # Conditional
            ("return %", None, 10),  # Conditional
            ("exit_type", Colors.LIGHT_YELLOW, 12),
            ("entry_reason", Colors.LIGHT_GREEN, 80),  # Wide for detailed AI signal
            ("exit_reason", Colors.LIGHT_GREEN, 60),   # Wide for detailed exit
        ]
        
        for col, (header, _, _) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)
        
        # Data rows
        row = 2
        for strategy in self.strategies:
            if strategy.trades.empty:
                continue
            
            for _, trade in strategy.trades.iterrows():
                col = 1
                
                # Identity columns (gray)
                cell = ws.cell(row=row, column=col, value=strategy.strategy_id)
                self._apply_fill(cell, Colors.LIGHT_GRAY)
                col += 1
                
                cell = ws.cell(row=row, column=col, value=trade.get('ticker', ''))
                self._apply_fill(cell, Colors.LIGHT_GRAY)
                col += 1
                
                # Direction column
                direction = trade.get('direction', 'LONG')
                cell = ws.cell(row=row, column=col, value=direction)
                self._apply_fill(cell, Colors.LIGHT_GRAY)
                col += 1
                
                # Timing columns (blue)
                cell = ws.cell(row=row, column=col, value=str(trade.get('entry_date', ''))[:10])
                self._apply_fill(cell, Colors.LIGHT_BLUE)
                col += 1
                
                cell = ws.cell(row=row, column=col, value=str(trade.get('exit_date', ''))[:10])
                self._apply_fill(cell, Colors.LIGHT_BLUE)
                col += 1
                
                # Hold days
                hold_days = trade.get('hold_days', 0)
                cell = ws.cell(row=row, column=col, value=hold_days)
                self._apply_fill(cell, Colors.LIGHT_BLUE)
                col += 1
                
                # Price columns (white)
                ws.cell(row=row, column=col, value=round(trade.get('entry_price', 0), 2))
                col += 1
                ws.cell(row=row, column=col, value=round(trade.get('exit_price', 0), 2))
                col += 1
                
                # Shares
                ws.cell(row=row, column=col, value=trade.get('shares', 0))
                col += 1
                
                # Results columns (conditional)
                pnl = trade.get('pnl', 0)
                cell = ws.cell(row=row, column=col, value=round(pnl, 2))
                self._apply_fill(cell, Colors.GREEN if pnl > 0 else Colors.RED if pnl < 0 else Colors.GRAY)
                col += 1
                
                ret = trade.get('return_pct', 0)
                cell = ws.cell(row=row, column=col, value=round(ret, 2))
                self._apply_fill(cell, Colors.GREEN if ret > 0 else Colors.RED if ret < 0 else Colors.GRAY)
                col += 1
                
                # Exit type (yellow) - simple categorization
                exit_type = trade.get('exit_type', trade.get('exit_reason', '')[:20] if trade.get('exit_reason') else '')
                cell = ws.cell(row=row, column=col, value=exit_type)
                self._apply_fill(cell, Colors.LIGHT_YELLOW)
                col += 1
                
                # DETAILED Entry Reason (green) - full AI signal explanation
                entry_reason = trade.get('entry_reason', 'signal')
                cell = ws.cell(row=row, column=col, value=entry_reason)
                self._apply_fill(cell, Colors.LIGHT_GREEN)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                col += 1
                
                # DETAILED Exit Reason (green) - full exit explanation
                exit_reason = trade.get('exit_reason', '')
                cell = ws.cell(row=row, column=col, value=exit_reason)
                self._apply_fill(cell, Colors.LIGHT_GREEN)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                col += 1
                
                row += 1
        
        # Adjust column widths based on header definitions
        for col, (_, _, width) in enumerate(headers, 1):
            col_letter = chr(64 + col) if col <= 26 else f"A{chr(64 + col - 26)}"
            ws.column_dimensions[col_letter].width = width
    
    def _create_per_stock_summary_sheet(self):
        """Create Sheet 4: Per_Stock_Summary."""
        ws = self.wb.create_sheet("Per_Stock_Summary")
        
        headers = [
            "strategy_id",
            "ticker",
            "total_trades",
            "win_rate %",
            "avg_return %",
            "total_pnl $",
            "max_drawdown %"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)
        
        row = 2
        for strategy in self.strategies:
            if strategy.trades.empty:
                continue
            
            # Group by ticker
            for ticker in strategy.trades['ticker'].unique():
                ticker_trades = strategy.trades[strategy.trades['ticker'] == ticker]
                
                total_trades = len(ticker_trades)
                wins = (ticker_trades['pnl'] > 0).sum()
                win_rate = (wins / total_trades * 100) if total_trades > 0 else 0
                avg_return = ticker_trades['return_pct'].mean() if 'return_pct' in ticker_trades else 0
                total_pnl = ticker_trades['pnl'].sum()
                
                # Calculate max drawdown for this ticker
                cumulative = ticker_trades['pnl'].cumsum()
                peak = cumulative.cummax()
                drawdown = ((cumulative - peak) / peak.replace(0, 1) * 100).min()
                
                ws.cell(row=row, column=1, value=strategy.strategy_id)
                ws.cell(row=row, column=2, value=ticker)
                ws.cell(row=row, column=3, value=total_trades)
                
                cell = ws.cell(row=row, column=4, value=round(win_rate, 2))
                self._apply_fill(cell, self._get_wr_color(win_rate))
                
                cell = ws.cell(row=row, column=5, value=round(avg_return, 2))
                self._apply_fill(cell, Colors.GREEN if avg_return > 0 else Colors.RED if avg_return < 0 else Colors.GRAY)
                
                cell = ws.cell(row=row, column=6, value=round(total_pnl, 2))
                self._apply_fill(cell, Colors.GREEN if total_pnl > 0 else Colors.RED if total_pnl < 0 else Colors.GRAY)
                
                cell = ws.cell(row=row, column=7, value=round(drawdown, 2))
                self._apply_fill(cell, self._get_dd_color(drawdown))
                
                row += 1
        
        # Adjust column widths
        for col in range(1, 8):
            ws.column_dimensions[chr(64 + col)].width = 15
    
    def _create_config_snapshot_sheet(self):
        """Create Sheet 5: Config_Snapshot."""
        ws = self.wb.create_sheet("Config_Snapshot")
        
        headers = ["strategy_id", "parameter", "value", "notes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)
        
        row = 2
        for strategy in self.strategies:
            for param, value in strategy.config.items():
                ws.cell(row=row, column=1, value=strategy.strategy_id)
                ws.cell(row=row, column=2, value=param)
                ws.cell(row=row, column=3, value=str(value))
                ws.cell(row=row, column=4, value="")
                row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 40
    
    def _create_how_to_read_sheet(self):
        """Create Sheet 6: How_To_Read."""
        ws = self.wb.create_sheet("How_To_Read")
        
        content = [
            ("METRIC DEFINITIONS", ""),
            ("", ""),
            ("Annual Growth % (CAGR)", "Compound Annual Growth Rate - average yearly return if gains were reinvested"),
            ("Worst Loss % (Max Drawdown)", "Largest peak-to-trough decline - measures worst case scenario"),
            ("Profit / Loss Ratio", "Average winning trade / Average losing trade - measures trade quality"),
            ("Win Rate %", "Percentage of trades that were profitable"),
            ("Total Trades", "Number of completed trades in the backtest period"),
            ("Best/Worst Year %", "Best and worst single-year performance"),
            ("Status", "PASS if all metrics meet minimum thresholds, FAIL otherwise"),
            ("", ""),
            ("COLOR RULES", ""),
            ("", ""),
            ("🟢 Green", "Good / Improvement / Target met"),
            ("🟡 Yellow", "Neutral / Acceptable / Within tolerance"),
            ("🔴 Red", "Bad / Risk / Below threshold"),
            ("⬜ Gray", "Informational only"),
            ("", ""),
            ("THRESHOLDS (Fixed Contract)", ""),
            ("", ""),
            ("CAGR", f"🟢 ≥ {Thresholds.CAGR_GOOD}%  |  🟡 {Thresholds.CAGR_ACCEPTABLE}-{Thresholds.CAGR_GOOD}%  |  🔴 < {Thresholds.CAGR_ACCEPTABLE}%"),
            ("Max Drawdown", f"🟢 ≤ {Thresholds.DD_GOOD}%  |  🟡 {Thresholds.DD_GOOD}-{Thresholds.DD_ACCEPTABLE}%  |  🔴 > {Thresholds.DD_ACCEPTABLE}%"),
            ("Profit Factor", f"🟢 ≥ {Thresholds.PF_GOOD}  |  🟡 {Thresholds.PF_ACCEPTABLE}-{Thresholds.PF_GOOD}  |  🔴 < {Thresholds.PF_ACCEPTABLE}"),
            ("Win Rate", f"🟢 ≥ {Thresholds.WR_GOOD}%  |  🟡 {Thresholds.WR_ACCEPTABLE}-{Thresholds.WR_GOOD}%  |  🔴 < {Thresholds.WR_ACCEPTABLE}%"),
            ("", ""),
            ("COMPARISON RULES", ""),
            ("", ""),
            ("Better", f"Metric improved by more than {Thresholds.COMPARISON_TOLERANCE}% vs baseline"),
            ("Same", f"Metric within ±{Thresholds.COMPARISON_TOLERANCE}% of baseline"),
            ("Worse", f"Metric declined by more than {Thresholds.COMPARISON_TOLERANCE}% vs baseline"),
            ("", ""),
            ("COLUMN GROUPS (Trades Sheet)", ""),
            ("", ""),
            ("Light Gray", "Identity columns (strategy_id, ticker, direction)"),
            ("Light Blue", "Timing columns (entry_date, exit_date, hold_days)"),
            ("White", "Price columns (entry_price, exit_price, shares)"),
            ("Light Yellow", "Exit type (stop_loss, take_profit, max_hold)"),
            ("Conditional", "Results columns (pnl $, return %) - green if positive, red if negative"),
            ("Light Green", "Signal Reason columns (entry_reason, exit_reason) - DETAILED explanations"),
            ("", ""),
            ("SIGNAL REASON FORMAT (entry_reason)", ""),
            ("", ""),
            ("AI Signal", "Primary driver - BULLISH/BEARISH with confidence level (HIGH/MEDIUM/LOW)"),
            ("Technical", "Supporting indicators - RSI, MACD, Bollinger Bands, Momentum, SMA trend, Volume"),
            ("Market", "Market context - SPY trend (above/below 200SMA), market RSI conditions"),
            ("", ""),
            ("EXIT REASON FORMAT (exit_reason)", ""),
            ("", ""),
            ("STOP LOSS", "Triggered when loss exceeds stop loss % - shows price, loss %, hold days"),
            ("TAKE PROFIT", "Triggered when gain exceeds take profit % - shows price, gain %, hold days"),
            ("MAX HOLD", "Triggered when max hold days reached - shows final P&L"),
            ("", ""),
            ("QUICK DECISION GUIDE", ""),
            ("", ""),
            ("✅ ACCEPT", "Status = PASS, all metrics green/yellow"),
            ("⚠️ REVIEW", "Status = PASS but some yellow metrics"),
            ("❌ REJECT", "Status = FAIL or any red metrics"),
        ]
        
        for row_idx, (col1, col2) in enumerate(content, 1):
            cell1 = ws.cell(row=row_idx, column=1, value=col1)
            cell2 = ws.cell(row=row_idx, column=2, value=col2)
            
            if col2 == "" and col1 != "":
                cell1.font = Font(bold=True, size=12)
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 80
    
    def generate(self, output_path: str):
        """Generate the complete Excel report."""
        if not self.strategies:
            raise ValueError("No strategies added to report")
        
        print(f"📊 Generating Excel report with {len(self.strategies)} strategy(ies)...")
        
        self._create_overview_sheet()
        self._create_comparison_sheet()
        self._create_trades_sheet()
        self._create_per_stock_summary_sheet()
        self._create_config_snapshot_sheet()
        self._create_how_to_read_sheet()
        
        self.wb.save(output_path)
        print(f"   ✅ Saved to: {output_path}")
        
        return output_path


def create_report_from_backtest(
    results: Dict,
    strategy_id: str,
    config: Dict = None,
    output_path: str = None
) -> str:
    """
    Create Excel report from backtest results dictionary.
    
    Args:
        results: Dictionary with keys like 'cagr_pct', 'max_drawdown_pct', etc.
        strategy_id: Unique identifier for this strategy
        config: Optional configuration dictionary
        output_path: Optional output path (auto-generated if not provided)
    
    Returns:
        Path to generated Excel file
    """
    if output_path is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = f'reports/strategy_report_{timestamp}.xlsx'
    
    # Create strategy result
    strategy = StrategyResult(
        strategy_id=strategy_id,
        cagr_pct=results.get('cagr_pct', 0),
        max_drawdown_pct=results.get('max_drawdown_pct', 0),
        profit_factor=results.get('profit_factor', 0),
        win_rate_pct=results.get('win_rate_pct', 0),
        total_trades=results.get('total_trades', 0),
        total_return_pct=results.get('total_return_pct', 0),
        best_year_pct=results.get('best_year_pct', 0),
        worst_year_pct=results.get('worst_year_pct', 0),
        trades=results.get('trades', pd.DataFrame()),
        config=config or {}
    )
    
    # Generate report
    writer = ExcelReportWriter()
    writer.add_strategy(strategy, is_baseline=True)
    
    return writer.generate(output_path)


if __name__ == "__main__":
    # Test with sample data
    sample_trades = pd.DataFrame({
        'ticker': ['AAPL', 'GOOGL', 'MSFT', 'AAPL'],
        'entry_date': ['2024-01-01', '2024-01-15', '2024-02-01', '2024-02-15'],
        'exit_date': ['2024-01-10', '2024-01-25', '2024-02-10', '2024-02-25'],
        'entry_price': [150.0, 140.0, 380.0, 155.0],
        'exit_price': [160.0, 135.0, 400.0, 165.0],
        'pnl': [1000, -500, 2000, 1000],
        'return_pct': [6.67, -3.57, 5.26, 6.45],
        'exit_reason': ['take_profit', 'stop_loss', 'take_profit', 'max_hold']
    })
    
    sample_results = {
        'cagr_pct': 28.5,
        'max_drawdown_pct': -12.3,
        'profit_factor': 1.8,
        'win_rate_pct': 58.0,
        'total_trades': 4,
        'total_return_pct': 35.0,
        'best_year_pct': 35.0,
        'worst_year_pct': 22.0,
        'trades': sample_trades
    }
    
    sample_config = {
        'position_size_pct': 15,
        'stop_loss_pct': 4,
        'take_profit_pct': 8,
        'max_hold_days': 10,
        'models': 'XGBoost + RandomForest + LightGBM'
    }
    
    output = create_report_from_backtest(
        results=sample_results,
        strategy_id='ai_ensemble_v1',
        config=sample_config,
        output_path='reports/test_strategy_report.xlsx'
    )
    
    print(f"\n✅ Test report generated: {output}")
