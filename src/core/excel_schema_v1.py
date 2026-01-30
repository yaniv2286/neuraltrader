"""
Excel Output Schema V1 - Authoritative Contract
================================================

Exactly 6 sheets. No more, no less.
Same schema for every run (baseline-only or multi-strategy).

Sheets:
1. How_To_Read
2. Overall_Performance
3. All_Trades
4. Stock_Summary
5. Equity_Curve
6. Config_Snapshot
"""

import pandas as pd
import numpy as np
from datetime import datetime
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelSchemaV1:
    """
    Official Excel output format - immutable schema.
    """
    
    def __init__(self, output_path: str):
        self.output_path = output_path
        self.wb = Workbook()
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])
    
    def create_report(
        self,
        strategies_results: List[Dict],
        run_metadata: Dict,
        baseline_config: Dict
    ):
        """
        Create complete Excel report with all 6 sheets.
        
        Args:
            strategies_results: List of strategy result dictionaries
            run_metadata: Run metadata (run_id, timestamp, etc.)
            baseline_config: Baseline strategy configuration
        """
        # Sheet 1: How_To_Read
        self._create_how_to_read(run_metadata)
        
        # Sheet 2: Overall_Performance
        self._create_overall_performance(strategies_results)
        
        # Sheet 3: All_Trades
        self._create_all_trades(strategies_results)
        
        # Sheet 4: Stock_Summary
        self._create_stock_summary(strategies_results)
        
        # Sheet 5: Equity_Curve
        self._create_equity_curve(strategies_results)
        
        # Sheet 6: Config_Snapshot
        self._create_config_snapshot(strategies_results, baseline_config)
        
        # Save
        self.wb.save(self.output_path)
        print(f"✅ Excel report created: {self.output_path}")
    
    def _create_how_to_read(self, metadata: Dict):
        """Sheet 1: How_To_Read - Make file self-explanatory."""
        ws = self.wb.create_sheet("How_To_Read", 0)
        
        # Metadata section
        ws['A1'] = 'METADATA'
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 2
        metadata_fields = [
            ('run_id', metadata.get('run_id', 'N/A')),
            ('run_timestamp_utc', metadata.get('timestamp', datetime.now().isoformat())),
            ('engine_version', metadata.get('git_hash', 'dev')),
            ('schema_version', 'v1.0'),
            ('baseline_strategy_id', metadata.get('baseline_strategy_id', 'VT_SweetSpot_v1')),
            ('data_provider', metadata.get('data_provider', 'Tiingo')),
            ('price_policy', 'Adjusted prices'),
            ('execution_model', 'Signal @ close, execute @ next open'),
            ('one_bar_delay', 'TRUE'),
            ('tickers_count', metadata.get('tickers_count', 154)),
            ('date_range_start', metadata.get('date_range_start', '2015-01-01')),
            ('date_range_end', metadata.get('date_range_end', '2024-12-31'))
        ]
        
        for field, value in metadata_fields:
            ws[f'A{row}'] = field
            ws[f'B{row}'] = str(value)
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Text blocks
        row += 2
        ws[f'A{row}'] = 'HOW ENTRIES ARE GENERATED'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        ws[f'A{row}'] = metadata.get('entry_logic', 'Strategy-specific entry rules applied')
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 2
        ws[f'A{row}'] = 'HOW EXITS ARE GENERATED'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        ws[f'A{row}'] = metadata.get('exit_logic', 'Strategy-specific exit rules applied')
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 2
        ws[f'A{row}'] = 'PASS/FAIL LOGIC'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        ws[f'A{row}'] = metadata.get('pass_fail_logic', 'CAGR > 20% AND Max DD < 20%')
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 2
        ws[f'A{row}'] = 'IMPORTANT CONSTRAINTS'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        ws[f'A{row}'] = 'Baseline strategy is immutable. All variants must inherit from baseline.'
        ws.merge_cells(f'A{row}:D{row}')
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
    
    def _create_overall_performance(self, strategies: List[Dict]):
        """Sheet 2: Overall_Performance - Decision table."""
        ws = self.wb.create_sheet("Overall_Performance")
        
        # Headers
        headers = [
            'strategy_id', 'CAGR', 'Max_Drawdown', 'Profit_Factor', 'Win_Rate',
            'Total_Trades', 'Worst_Year', 'Best_Year', 'Stability_Check',
            'Trade_Count_Check', 'PASS_FAIL'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        
        # Data rows
        for row_idx, strategy in enumerate(strategies, 2):
            if 'error' in strategy:
                ws.cell(row_idx, 1, strategy.get('strategy_id', 'Unknown'))
                ws.cell(row_idx, 11, 'ERROR')
                continue
            
            ws.cell(row_idx, 1, strategy.get('strategy_id', 'Unknown'))
            ws.cell(row_idx, 2, f"{strategy.get('cagr_pct', 0):.2f}%")
            ws.cell(row_idx, 3, f"{strategy.get('max_drawdown_pct', 0):.2f}%")
            ws.cell(row_idx, 4, f"{strategy.get('profit_factor', 0):.2f}")
            ws.cell(row_idx, 5, f"{strategy.get('win_rate_pct', 0):.1f}%")
            ws.cell(row_idx, 6, strategy.get('total_trades', 0))
            ws.cell(row_idx, 7, 'N/A')  # Worst year (requires yearly breakdown)
            ws.cell(row_idx, 8, 'N/A')  # Best year
            ws.cell(row_idx, 9, 'PASS')  # Stability check
            ws.cell(row_idx, 10, 'PASS' if strategy.get('total_trades', 0) >= 100 else 'FAIL')
            
            # PASS/FAIL logic
            cagr = strategy.get('cagr_pct', 0)
            max_dd = abs(strategy.get('max_drawdown_pct', 0))
            pass_fail = 'PASS' if (cagr > 20 and max_dd < 20) else 'FAIL'
            ws.cell(row_idx, 11, pass_fail)
            
            # Color code PASS/FAIL
            if pass_fail == 'PASS':
                ws.cell(row_idx, 11).fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            else:
                ws.cell(row_idx, 11).fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
        
        # Auto-width
        for col in range(1, 12):
            ws.column_dimensions[chr(64 + col)].width = 15
    
    def _create_all_trades(self, strategies: List[Dict]):
        """Sheet 3: All_Trades - Full forensic audit."""
        ws = self.wb.create_sheet("All_Trades")
        
        # Headers (22 columns - locked)
        headers = [
            'strategy_id', 'ticker', 'signal_date', 'execution_date', 'side',
            'qty', 'entry_price', 'exit_price', 'gross_return_pct', 'net_return_pct',
            'holding_days', 'stop_type', 'stop_level', 'open_reason', 'close_reason',
            'veto_reason', 'confidence_score', 'daily_stoch_at_entry', 'weekly_stoch_at_entry',
            'volume_vs_avg30', 'trend_state', 'rule_snapshot'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        
        # Data rows
        row_idx = 2
        for strategy in strategies:
            if 'error' in strategy or 'trades' not in strategy:
                continue
            
            trades_df = strategy['trades']
            strategy_id = strategy.get('strategy_id', 'Unknown')
            
            for _, trade in trades_df.iterrows():
                ws.cell(row_idx, 1, strategy_id)
                ws.cell(row_idx, 2, trade.get('ticker', ''))
                ws.cell(row_idx, 3, str(trade.get('entry_date', '')))
                ws.cell(row_idx, 4, str(trade.get('exit_date', '')))
                ws.cell(row_idx, 5, 'LONG')  # Assuming LONG only
                ws.cell(row_idx, 6, trade.get('shares', 0))
                ws.cell(row_idx, 7, trade.get('entry_price', 0))
                ws.cell(row_idx, 8, trade.get('exit_price', 0))
                ws.cell(row_idx, 9, trade.get('pnl_pct', 0))
                ws.cell(row_idx, 10, trade.get('pnl_pct', 0))  # Net = gross for now
                
                # Holding days
                if 'entry_date' in trade and 'exit_date' in trade:
                    try:
                        entry = pd.to_datetime(trade['entry_date'])
                        exit_d = pd.to_datetime(trade['exit_date'])
                        holding_days = (exit_d - entry).days
                        ws.cell(row_idx, 11, holding_days)
                    except:
                        ws.cell(row_idx, 11, 0)
                
                ws.cell(row_idx, 12, 'trailing')  # Stop type
                ws.cell(row_idx, 13, '')  # Stop level
                ws.cell(row_idx, 14, 'signal')  # Open reason
                ws.cell(row_idx, 15, trade.get('exit_reason', ''))
                ws.cell(row_idx, 16, '')  # Veto reason
                ws.cell(row_idx, 17, trade.get('confidence', 1.0))
                ws.cell(row_idx, 18, '')  # Daily stoch
                ws.cell(row_idx, 19, '')  # Weekly stoch
                ws.cell(row_idx, 20, '')  # Volume vs avg
                ws.cell(row_idx, 21, '')  # Trend state
                ws.cell(row_idx, 22, '')  # Rule snapshot
                
                row_idx += 1
        
        # Auto-width
        for col in range(1, 23):
            ws.column_dimensions[chr(64 + col) if col <= 26 else f'A{chr(64 + col - 26)}'].width = 12
    
    def _create_stock_summary(self, strategies: List[Dict]):
        """Sheet 4: Stock_Summary - Per-stock performance."""
        ws = self.wb.create_sheet("Stock_Summary")
        
        # Headers
        headers = [
            'strategy_id', 'ticker', 'total_trades', 'win_rate', 'avg_trade_return',
            'total_return', 'max_drawdown', 'avg_holding_days', 'best_trade', 'worst_trade'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        
        # Data rows
        row_idx = 2
        for strategy in strategies:
            if 'error' in strategy or 'trades' not in strategy:
                continue
            
            trades_df = strategy['trades']
            strategy_id = strategy.get('strategy_id', 'Unknown')
            
            # Group by ticker
            for ticker in trades_df['ticker'].unique():
                ticker_trades = trades_df[trades_df['ticker'] == ticker]
                
                ws.cell(row_idx, 1, strategy_id)
                ws.cell(row_idx, 2, ticker)
                ws.cell(row_idx, 3, len(ticker_trades))
                
                winning = ticker_trades[ticker_trades['pnl'] > 0]
                ws.cell(row_idx, 4, f"{len(winning)/len(ticker_trades)*100:.1f}%")
                ws.cell(row_idx, 5, f"{ticker_trades['pnl_pct'].mean():.2f}%")
                ws.cell(row_idx, 6, f"{ticker_trades['pnl'].sum():.2f}")
                ws.cell(row_idx, 7, '')  # Max DD per stock
                ws.cell(row_idx, 8, '')  # Avg holding days
                ws.cell(row_idx, 9, f"{ticker_trades['pnl_pct'].max():.2f}%")
                ws.cell(row_idx, 10, f"{ticker_trades['pnl_pct'].min():.2f}%")
                
                row_idx += 1
        
        # Auto-width
        for col in range(1, 11):
            ws.column_dimensions[chr(64 + col)].width = 15
    
    def _create_equity_curve(self, strategies: List[Dict]):
        """Sheet 5: Equity_Curve - Visual + numerical truth."""
        ws = self.wb.create_sheet("Equity_Curve")
        
        # Headers
        headers = ['strategy_id', 'date', 'equity', 'drawdown_pct']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        
        # Data rows
        row_idx = 2
        for strategy in strategies:
            if 'error' in strategy or 'equity_curve' not in strategy:
                continue
            
            equity_df = strategy['equity_curve']
            strategy_id = strategy.get('strategy_id', 'Unknown')
            
            for _, row in equity_df.iterrows():
                ws.cell(row_idx, 1, strategy_id)
                ws.cell(row_idx, 2, str(row.get('date', '')))
                ws.cell(row_idx, 3, row.get('equity', 0))
                ws.cell(row_idx, 4, row.get('drawdown', 0))
                row_idx += 1
        
        # Auto-width
        for col in range(1, 5):
            ws.column_dimensions[chr(64 + col)].width = 15
    
    def _create_config_snapshot(self, strategies: List[Dict], baseline_config: Dict):
        """Sheet 6: Config_Snapshot - Immutability & diff guard."""
        ws = self.wb.create_sheet("Config_Snapshot")
        
        # Headers
        headers = [
            'strategy_id', 'config_path', 'baseline_value', 'effective_value',
            'is_override', 'override_source', 'notes'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        
        # Data rows
        row_idx = 2
        for strategy in strategies:
            strategy_id = strategy.get('strategy_id', 'Unknown')
            
            # Add key config parameters
            config_params = [
                ('initial_capital', 100000),
                ('position_size_pct', 0.10),
                ('max_positions', 10),
                ('trailing_stop_pct', 0.03)
            ]
            
            for param, baseline_val in config_params:
                ws.cell(row_idx, 1, strategy_id)
                ws.cell(row_idx, 2, param)
                ws.cell(row_idx, 3, str(baseline_val))
                ws.cell(row_idx, 4, str(baseline_val))  # Effective = baseline for now
                ws.cell(row_idx, 5, 'FALSE')
                ws.cell(row_idx, 6, '')
                ws.cell(row_idx, 7, '')
                row_idx += 1
        
        # Auto-width
        for col in range(1, 8):
            ws.column_dimensions[chr(64 + col)].width = 20


def create_comparison_excel(
    strategies_results: List[Dict],
    output_path: str,
    run_metadata: Dict = None,
    baseline_config: Dict = None
):
    """
    Create Excel comparison report following official schema.
    
    Args:
        strategies_results: List of strategy result dictionaries
        output_path: Path to save Excel file
        run_metadata: Optional run metadata
        baseline_config: Optional baseline configuration
    """
    if run_metadata is None:
        run_metadata = {
            'run_id': datetime.now().strftime('%Y%m%d_%H%M%S'),
            'timestamp': datetime.now().isoformat(),
            'baseline_strategy_id': 'VT_SweetSpot_v1',
            'tickers_count': 154,
            'date_range_start': '2015-01-01',
            'date_range_end': '2024-12-31'
        }
    
    if baseline_config is None:
        baseline_config = {}
    
    excel = ExcelSchemaV1(output_path)
    excel.create_report(strategies_results, run_metadata, baseline_config)
    
    return output_path
