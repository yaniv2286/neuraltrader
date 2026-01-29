#!/usr/bin/env python3
"""
Generic Report Generation Module for NeuralTrader
Creates comprehensive Excel reports and visualizations
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import matplotlib.pyplot as plt
import seaborn as sns
from typing import Dict, List, Any, Optional
import json

# Excel libraries
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.chart import BarChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not available, Excel reports will be limited")

class ReportGenerator:
    """Generic report generation class"""
    
    def __init__(self, output_dir: str = None):
        self.output_dir = output_dir or "reports"
        Path(self.output_dir).mkdir(exist_ok=True)
        
    def generate_excel_report(self, results_df: pd.DataFrame, 
                            summary: Dict = None,
                            report_name: str = "analysis_report") -> str:
        """Generate comprehensive Excel report"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{report_name}_{timestamp}.xlsx"
        filepath = Path(self.output_dir) / filename
        
        if not EXCEL_AVAILABLE:
            # Fallback to CSV
            csv_filepath = filepath.with_suffix('.csv')
            results_df.to_csv(csv_filepath, index=False)
            return str(csv_filepath)
        
        # Create Excel workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # 1. Summary Sheet
        self._create_summary_sheet(wb, summary)
        
        # 2. Results Sheet
        self._create_results_sheet(wb, results_df)
        
        # 3. Top Performers Sheet
        self._create_top_performers_sheet(wb, results_df)
        
        # 4. Analysis Sheet
        self._create_analysis_sheet(wb, results_df)
        
        # 5. Charts Sheet
        self._create_charts_sheet(wb, results_df)
        
        # Save workbook
        wb.save(filepath)
        return str(filepath)
    
    def _create_summary_sheet(self, wb, summary: Dict):
        """Create summary sheet"""
        ws = wb.create_sheet("Summary")
        
        # Title
        ws['A1'] = "NEURALTRADER ANALYSIS REPORT"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:B1')
        
        # Date
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'].font = Font(bold=True)
        
        if summary:
            row = 5
            for key, value in summary.items():
                if isinstance(value, list):
                    ws[f'A{row}'] = key.replace('_', ' ').title()
                    ws[f'B{row}'] = len(value)
                    row += 1
                elif isinstance(value, dict):
                    ws[f'A{row}'] = key.replace('_', ' ').title()
                    ws[f'B{row}'] = json.dumps(value, indent=2)
                    row += 2
                else:
                    ws[f'A{row}'] = key.replace('_', ' ').title()
                    ws[f'B{row}'] = str(value)
                    row += 1
    
    def _create_results_sheet(self, wb, results_df: pd.DataFrame):
        """Create detailed results sheet"""
        ws = wb.create_sheet("Results")
        
        # Headers
        headers = results_df.columns.tolist()
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        for row_idx, row in results_df.iterrows():
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx + 2, column=col_idx, value=value)
    
    def _create_top_performers_sheet(self, wb, results_df: pd.DataFrame):
        """Create top performers sheet"""
        ws = wb.create_sheet("Top Performers")
        
        # Filter successful results
        successful = results_df[results_df['error'].isna()]
        
        if len(successful) > 0:
            # Top 10 by R²
            top_r2 = successful.nlargest(10, 'test_r2')
            
            # Headers
            headers = ['Rank', 'Ticker', 'Test R²', 'Test Direction', 'Generalization Gap']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Data
            for idx, (_, row) in enumerate(top_r2.iterrows(), 1):
                ws.cell(row=idx + 1, column=1, value=idx)
                ws.cell(row=idx + 1, column=2, value=row['ticker'])
                ws.cell(row=idx + 1, column=3, value=row['test_r2'])
                ws.cell(row=idx + 1, column=4, value=row['test_dir'])
                ws.cell(row=idx + 1, column=5, value=row['gen_gap'])
    
    def _create_analysis_sheet(self, wb, results_df: pd.DataFrame):
        """Create analysis sheet with statistics"""
        ws = wb.create_sheet("Analysis")
        
        # Filter successful results
        successful = results_df[results_df['error'].isna()]
        
        if len(successful) > 0:
            # Statistics
            stats = {
                'Mean Test R²': successful['test_r2'].mean(),
                'Std Test R²': successful['test_r2'].std(),
                'Mean Test Direction': successful['test_dir'].mean(),
                'Std Test Direction': successful['test_dir'].std(),
                'Mean Generalization Gap': successful['gen_gap'].mean(),
                'Good Models (%)': (successful['is_good'].sum() / len(successful) * 100)
            }
            
            row = 1
            for stat, value in stats.items():
                ws.cell(row=row, column=1, value=stat)
                ws.cell(row=row, column=2, value=value)
                row += 1
    
    def _create_charts_sheet(self, wb, results_df: pd.DataFrame):
        """Create charts sheet"""
        ws = wb.create_sheet("Charts")
        
        # Filter successful results
        successful = results_df[results_df['error'].isna()]
        
        if len(successful) > 0:
            # Create a simple bar chart for top performers
            top_10 = successful.nlargest(10, 'test_r2')
            
            # Chart data
            for idx, (_, row) in enumerate(top_10.iterrows(), 1):
                ws.cell(row=idx, column=1, value=row['ticker'])
                ws.cell(row=idx, column=2, value=row['test_r2'])
            
            # Create chart
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Top 10 Performers by Test R²"
            chart.y_axis.title = 'Test R²'
            chart.x_axis.title = 'Ticker'
            
            data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=len(top_10))
            labels = Reference(ws, min_col=1, min_row=1, max_row=len(top_10))
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(labels)
            
            ws.add_chart(chart, "E2")
    
    def generate_visualizations(self, results_df: pd.DataFrame, 
                              save_individual: bool = True) -> List[str]:
        """Generate visualization charts"""
        successful = results_df[results_df['error'].isna()]
        
        if len(successful) == 0:
            return []
        
        # Set style
        plt.style.use('seaborn-v0_8')
        sns.set_palette("husl")
        
        charts = []
        
        # 1. R² Distribution
        plt.figure(figsize=(10, 6))
        sns.histplot(successful['test_r2'], bins=30, kde=True)
        plt.title('Distribution of Test R² Scores')
        plt.xlabel('Test R²')
        plt.ylabel('Frequency')
        chart_path = self._save_chart('r2_distribution.png')
        charts.append(chart_path)
        
        # 2. Direction Accuracy Distribution
        plt.figure(figsize=(10, 6))
        sns.histplot(successful['test_dir'], bins=30, kde=True)
        plt.title('Distribution of Direction Accuracy')
        plt.xlabel('Direction Accuracy (%)')
        plt.ylabel('Frequency')
        chart_path = self._save_chart('direction_accuracy.png')
        charts.append(chart_path)
        
        # 3. R² vs Direction Accuracy Scatter
        plt.figure(figsize=(10, 6))
        sns.scatterplot(data=successful, x='test_r2', y='test_dir', alpha=0.6)
        plt.title('R² vs Direction Accuracy')
        plt.xlabel('Test R²')
        plt.ylabel('Direction Accuracy (%)')
        chart_path = self._save_chart('r2_vs_direction.png')
        charts.append(chart_path)
        
        # 4. Top Performers Bar Chart
        top_10 = successful.nlargest(10, 'test_r2')
        plt.figure(figsize=(12, 6))
        sns.barplot(data=top_10, x='ticker', y='test_r2')
        plt.title('Top 10 Performers by Test R²')
        plt.xlabel('Ticker')
        plt.ylabel('Test R²')
        plt.xticks(rotation=45)
        chart_path = self._save_chart('top_performers.png')
        charts.append(chart_path)
        
        # 5. Generalization Gap Distribution
        plt.figure(figsize=(10, 6))
        sns.histplot(successful['gen_gap'], bins=30, kde=True)
        plt.title('Distribution of Generalization Gap')
        plt.xlabel('Generalization Gap')
        plt.ylabel('Frequency')
        chart_path = self._save_chart('gen_gap_distribution.png')
        charts.append(chart_path)
        
        return charts
    
    def _save_chart(self, filename: str) -> str:
        """Save chart to file"""
        filepath = Path(self.output_dir) / filename
        plt.savefig(filepath, dpi=300, bbox_inches='tight')
        plt.close()
        return str(filepath)
    
    def generate_full_report(self, results_df: pd.DataFrame, 
                           summary: Dict = None,
                           report_name: str = "comprehensive_analysis") -> Dict[str, str]:
        """Generate full report with Excel and visualizations"""
        print("📊 Generating comprehensive report...")
        
        # Generate Excel report
        excel_file = self.generate_excel_report(results_df, summary, report_name)
        print(f"   ✅ Excel report: {excel_file}")
        
        # Generate visualizations
        charts = self.generate_visualizations(results_df)
        print(f"   ✅ Generated {len(charts)} charts")
        
        return {
            "excel_report": excel_file,
            "charts": charts,
            "report_name": report_name
        }

def main():
    """Command line interface for report generation"""
    import argparse
    
    parser = argparse.ArgumentParser(description='NeuralTrader Report Generator')
    parser.add_argument('--results-file', type=str, required=True,
                       help='CSV file with test results')
    parser.add_argument('--summary-file', type=str,
                       help='JSON file with summary statistics')
    parser.add_argument('--report-name', type=str, default='analysis_report',
                       help='Name for the report files')
    parser.add_argument('--output-dir', type=str, default='reports',
                       help='Output directory for reports')
    
    args = parser.parse_args()
    
    # Load results
    results_df = pd.read_csv(args.results_file)
    
    # Load summary if provided
    summary = None
    if args.summary_file:
        with open(args.summary_file, 'r') as f:
            summary = json.load(f)
    
    # Generate report
    generator = ReportGenerator(args.output_dir)
    report = generator.generate_full_report(results_df, summary, args.report_name)
    
    print(f"\n🎉 Report generation complete!")
    print(f"📁 Excel: {report['excel_report']}")
    print(f"📊 Charts: {len(report['charts'])} files")

if __name__ == "__main__":
    main()
